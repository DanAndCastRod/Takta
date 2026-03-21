"""
Assets Router — CRUD + Tree + Breadcrumbs for the Asset hierarchy.
"""
from io import BytesIO
from typing import List, Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select
from pydantic import BaseModel
import uuid as uuid_lib

from ...db import get_session
from ...models import Asset
from ...core.auth import get_current_user, require_role, CurrentUser

router = APIRouter(
    prefix="/api/assets",
    tags=["assets"],
    dependencies=[Depends(get_current_user)],
)


# --- Schemas ---

class AssetCreate(BaseModel):
    """Schema for creating a new asset."""
    name: str
    type: str  # Sede, Planta, Area, Linea, Maquina, Puesto, Componente
    description: Optional[str] = None
    parent_id: Optional[str] = None  # UUID as string


class AssetRead(BaseModel):
    """Schema for reading an asset (flat)."""
    id: str
    name: str
    type: str
    description: Optional[str] = None
    parent_id: Optional[str] = None


class AssetTreeNode(BaseModel):
    """Schema for recursive tree node."""
    id: str
    name: str
    type: str
    description: Optional[str] = None
    children: List["AssetTreeNode"] = []


class BulkAssetImportResponse(BaseModel):
    created: int
    updated: int
    errors_count: int
    errors: List[str]


# --- Helpers ---

def _build_tree_node(asset: Asset) -> dict:
    """Recursively build a tree node dict from an Asset with eager-loaded children."""
    return {
        "id": str(asset.id),
        "name": asset.name,
        "type": asset.type,
        "description": asset.description,
        "children": [_build_tree_node(child) for child in (asset.children or [])]
    }


def _detect_cycle(session: Session, asset_id: uuid_lib.UUID, proposed_parent_id: uuid_lib.UUID) -> bool:
    """
    Check if setting proposed_parent_id as parent of asset_id would create a cycle.
    Walks up from proposed_parent_id to root, checking if we encounter asset_id.
    """
    current_id = proposed_parent_id
    visited = set()
    
    while current_id is not None:
        if current_id == asset_id:
            return True  # Cycle detected!
        if current_id in visited:
            return True  # Already in a cycle
        visited.add(current_id)
        
        parent = session.get(Asset, current_id)
        if parent is None:
            break
        current_id = parent.parent_id
    
    return False


def _get_breadcrumbs(session: Session, asset_id: uuid_lib.UUID) -> List[dict]:
    """
    Walk up the asset tree from asset_id to root, returning the full path.
    Returns list from root to current asset.
    """
    path = []
    current_id = asset_id
    
    while current_id is not None:
        asset = session.get(Asset, current_id)
        if asset is None:
            break
        path.append({
            "id": str(asset.id),
            "name": asset.name,
            "type": asset.type,
        })
        current_id = asset.parent_id
    
    path.reverse()  # Root first
    return path


def _norm_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _asset_code(asset: Asset) -> str:
    return f"A-{str(asset.id)[:8]}"


def _workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _build_assets_template_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "assets"
    sheet.append(["asset_id", "name", "type", "description", "code", "parent_code"])
    sheet.append(["", "Sede Central", "Sede", "Raiz", "SEDE", ""])
    sheet.append(["", "Planta Principal", "Planta", "Hija de sede", "PLANTA", "SEDE"])
    sheet.append(["", "Linea 1", "Linea", "Hija de planta", "LINEA-1", "PLANTA"])

    notes = workbook.create_sheet("instrucciones")
    notes.append(["Reglas de carga masiva de Activos"])
    notes.append(["1) name y type son obligatorios."])
    notes.append(["2) code identifica la fila dentro del archivo para construir jerarquia."])
    notes.append(["3) parent_code puede apuntar a code o a un UUID existente."])
    notes.append(["4) asset_id opcional: si se llena, actualiza el activo existente."])
    notes.append(["5) si asset_id esta vacio, se crea un activo nuevo."])
    notes.append(["6) parent_code vacio crea nodo raiz (o conserva parent en updates)."])
    return workbook


def _build_assets_export_workbook(session: Session) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "assets"
    sheet.append(["asset_id", "name", "type", "description", "code", "parent_code"])

    assets = session.exec(select(Asset)).all()
    codes = {asset.id: _asset_code(asset) for asset in assets}

    for asset in assets:
        sheet.append(
            [
                str(asset.id),
                asset.name,
                asset.type,
                asset.description,
                codes.get(asset.id),
                codes.get(asset.parent_id) if asset.parent_id else None,
            ]
        )
    return workbook


# --- Endpoints ---

@router.get("", response_model=List[AssetRead])
def list_assets(
    offset: int = 0,
    limit: int = Query(default=100, le=1000),
    type: Optional[str] = Query(default=None, description="Filter by asset type"),
    session: Session = Depends(get_session),
):
    """List all assets (flat, paginated). Optionally filter by type."""
    statement = select(Asset)
    if type:
        statement = statement.where(Asset.type == type)
    statement = statement.offset(offset).limit(limit)
    
    assets = session.exec(statement).all()
    return [
        AssetRead(
            id=str(a.id), name=a.name, type=a.type,
            description=a.description,
            parent_id=str(a.parent_id) if a.parent_id else None
        )
        for a in assets
    ]


@router.post("", response_model=AssetRead, status_code=201)
def create_asset(
    payload: AssetCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    """
    Create a new asset node.
    Validates that the parent exists and no cycle would be created.
    """
    parent_uuid = None
    if payload.parent_id:
        parent_uuid = uuid_lib.UUID(payload.parent_id)
        parent = session.get(Asset, parent_uuid)
        if not parent:
            raise HTTPException(status_code=404, detail=f"Parent asset '{payload.parent_id}' not found")
    
    asset = Asset(
        name=payload.name,
        type=payload.type,
        description=payload.description,
        parent_id=parent_uuid,
    )
    session.add(asset)
    session.commit()
    session.refresh(asset)
    
    return AssetRead(
        id=str(asset.id), name=asset.name, type=asset.type,
        description=asset.description,
        parent_id=str(asset.parent_id) if asset.parent_id else None
    )


@router.get("/tree")
def get_asset_tree(session: Session = Depends(get_session)):
    """
    Return the complete asset hierarchy as nested JSON.
    Starts from root nodes (no parent) and recursively includes children.
    """
    # Fetch root nodes
    roots = session.exec(select(Asset).where(Asset.parent_id == None)).all()
    
    # Build recursive tree
    tree = [_build_tree_node(root) for root in roots]
    return tree


@router.get("/xlsx/template")
def download_assets_template():
    workbook = _build_assets_template_workbook()
    return _workbook_response(workbook, "takta_assets_template.xlsx")


@router.get("/xlsx/export")
def download_assets_export(session: Session = Depends(get_session)):
    workbook = _build_assets_export_workbook(session)
    return _workbook_response(workbook, "takta_assets_export.xlsx")


@router.post("/xlsx/import", response_model=BulkAssetImportResponse)
def import_assets_xlsx(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file.")

    try:
        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLSX file.") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    created = 0
    updated = 0
    errors: List[str] = []
    code_map: dict[str, Asset] = {}
    pending_parents: List[tuple[Asset, str, int]] = []

    for row_idx, row in enumerate(rows, start=2):
        asset_id_raw = _norm_text(row[0] if len(row) > 0 else None)
        name = _norm_text(row[1] if len(row) > 1 else None)
        asset_type = _norm_text(row[2] if len(row) > 2 else None)
        description = _norm_text(row[3] if len(row) > 3 else None)
        code = _norm_text(row[4] if len(row) > 4 else None)
        parent_code = _norm_text(row[5] if len(row) > 5 else None)

        if not any([asset_id_raw, name, asset_type, description, code, parent_code]):
            continue

        if not name or not asset_type:
            errors.append(f"Row {row_idx}: name and type are required.")
            continue

        asset: Optional[Asset] = None
        if asset_id_raw:
            try:
                asset_uuid = uuid_lib.UUID(asset_id_raw)
            except ValueError:
                errors.append(f"Row {row_idx}: invalid asset_id UUID '{asset_id_raw}'.")
                continue

            asset = session.get(Asset, asset_uuid)
            if not asset:
                errors.append(f"Row {row_idx}: asset_id '{asset_id_raw}' not found.")
                continue

            asset.name = name
            asset.type = asset_type
            asset.description = description
            updated += 1
        else:
            asset = Asset(
                name=name,
                type=asset_type,
                description=description,
                parent_id=None,
            )
            session.add(asset)
            session.flush()
            created += 1

        session.add(asset)

        resolved_code = (code or str(asset.id)).strip().lower()
        code_map[resolved_code] = asset
        code_map[str(asset.id).lower()] = asset
        if parent_code:
            pending_parents.append((asset, parent_code.strip(), row_idx))

    for asset, parent_ref, row_idx in pending_parents:
        parent = code_map.get(parent_ref.lower())
        if not parent:
            try:
                parent_uuid = uuid_lib.UUID(parent_ref)
                parent = session.get(Asset, parent_uuid)
            except ValueError:
                parent = None

        if not parent:
            errors.append(f"Row {row_idx}: parent '{parent_ref}' not found.")
            continue

        if parent.id == asset.id:
            errors.append(f"Row {row_idx}: parent cannot be the same asset.")
            continue

        if _detect_cycle(session, asset.id, parent.id):
            errors.append(f"Row {row_idx}: parent '{parent_ref}' creates cycle.")
            continue

        asset.parent_id = parent.id
        session.add(asset)

    session.commit()
    return BulkAssetImportResponse(
        created=created,
        updated=updated,
        errors_count=len(errors),
        errors=errors,
    )


@router.get("/{asset_id}")
def get_asset(asset_id: str, session: Session = Depends(get_session)):
    """Get a specific asset by ID with its direct children."""
    asset = session.get(Asset, uuid_lib.UUID(asset_id))
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    return {
        "id": str(asset.id),
        "name": asset.name,
        "type": asset.type,
        "description": asset.description,
        "parent_id": str(asset.parent_id) if asset.parent_id else None,
        "children": [
            {
                "id": str(c.id),
                "name": c.name,
                "type": c.type,
                "description": c.description,
            }
            for c in (asset.children or [])
        ]
    }


@router.get("/{asset_id}/context")
def get_asset_context(asset_id: str, session: Session = Depends(get_session)):
    """
    Return the full breadcrumb path from root to this asset.
    Example: ["Sede Pereira", "Planta Beneficio", "Área Evisceración", "Línea 1"]
    """
    asset = session.get(Asset, uuid_lib.UUID(asset_id))
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    breadcrumbs = _get_breadcrumbs(session, uuid_lib.UUID(asset_id))
    
    return {
        "asset_id": asset_id,
        "asset_name": asset.name,
        "breadcrumbs": breadcrumbs,
        "depth": len(breadcrumbs),
    }


@router.patch("/{asset_id}")
def update_asset(
    asset_id: str,
    payload: AssetCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    """Update an asset's name, type, description, or parent. Validates against cycles."""
    asset = session.get(Asset, uuid_lib.UUID(asset_id))
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    # Check for cycle if parent is changing
    if payload.parent_id:
        new_parent_uuid = uuid_lib.UUID(payload.parent_id)
        if _detect_cycle(session, asset.id, new_parent_uuid):
            raise HTTPException(
                status_code=400,
                detail="Cannot set this parent: would create a circular reference"
            )
        asset.parent_id = new_parent_uuid
    
    asset.name = payload.name
    asset.type = payload.type
    asset.description = payload.description
    
    session.add(asset)
    session.commit()
    session.refresh(asset)
    
    return AssetRead(
        id=str(asset.id), name=asset.name, type=asset.type,
        description=asset.description,
        parent_id=str(asset.parent_id) if asset.parent_id else None
    )


@router.delete("/{asset_id}", status_code=204)
def delete_asset(
    asset_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    """Delete an asset. Fails if it has children (prevent orphans)."""
    asset = session.get(Asset, uuid_lib.UUID(asset_id))
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if asset.children:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: asset has {len(asset.children)} children. Delete or reassign them first."
        )
    
    session.delete(asset)
    session.commit()

