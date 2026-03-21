"""
Engineering API:
- Activities CRUD
- References (SKU) CRUD with measurement fields
- Standards (Asset + Activity + SKU) CRUD
- Time studies (chronometer) with association to asset/SKU/standard
- Apply chronometer results to standards
- XLSX template/export/import for bulk loading
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, List, Optional
import uuid as uuid_lib

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlmodel import Session, and_, select

from ..core.auth import CurrentUser, get_current_user, require_role
from ..db import get_session
from ..models import (
    Asset,
    ProcessStandard,
    ProductReference,
    StandardActivity,
    TimeStudy,
    TimingElement,
    TimingLap,
    TimingSession,
)


router = APIRouter(
    prefix="/api/engineering",
    tags=["Engineering"],
    dependencies=[Depends(get_current_user)],
)


def _to_uuid(raw: Optional[str], field_name: str) -> Optional[uuid_lib.UUID]:
    if raw in (None, ""):
        return None
    try:
        return uuid_lib.UUID(raw)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=f"Invalid UUID for '{field_name}'") from exc


def _bool_value(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "si"}


def _text_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    txt = str(value).strip()
    return txt if txt else None


def _float_value(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _build_standard_read(session: Session, standard: ProcessStandard) -> "StandardRead":
    asset = session.get(Asset, standard.asset_id)
    activity = session.get(StandardActivity, standard.activity_id)
    reference = (
        session.get(ProductReference, standard.product_reference_id)
        if standard.product_reference_id
        else None
    )
    return StandardRead(
        id=str(standard.id),
        asset_id=str(standard.asset_id),
        activity_id=str(standard.activity_id),
        product_reference_id=str(standard.product_reference_id) if standard.product_reference_id else None,
        standard_time_minutes=standard.standard_time_minutes,
        frequency=standard.frequency,
        is_active=standard.is_active,
        capacity_unit=standard.capacity_unit,
        asset_name=asset.name if asset else None,
        activity_name=activity.name if activity else None,
        reference_code=reference.code if reference else None,
    )


def _study_linked_asset_id(study: TimeStudy, standard: Optional[ProcessStandard]) -> Optional[uuid_lib.UUID]:
    return study.asset_id or (standard.asset_id if standard else None)


def _study_linked_reference_id(
    study: TimeStudy, standard: Optional[ProcessStandard]
) -> Optional[uuid_lib.UUID]:
    return study.product_reference_id or (standard.product_reference_id if standard else None)


class ActivityCreate(BaseModel):
    name: str
    type: str
    is_value_added: bool = False


class ActivityUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    is_value_added: Optional[bool] = None


class ActivityRead(BaseModel):
    id: str
    name: str
    type: str
    is_value_added: bool


class ReferenceCreate(BaseModel):
    code: str
    description: str
    family: str
    uom: Optional[str] = None
    packaging_uom: Optional[str] = None


class ReferenceUpdate(BaseModel):
    code: Optional[str] = None
    description: Optional[str] = None
    family: Optional[str] = None
    uom: Optional[str] = None
    packaging_uom: Optional[str] = None


class ReferenceRead(BaseModel):
    id: str
    code: str
    description: str
    family: str
    uom: Optional[str] = None
    packaging_uom: Optional[str] = None


class StandardCreate(BaseModel):
    asset_id: str
    activity_id: str
    product_reference_id: Optional[str] = None
    standard_time_minutes: Optional[float] = None
    frequency: Optional[str] = None
    capacity_unit: Optional[str] = None


class StandardUpdate(BaseModel):
    is_active: Optional[bool] = None
    standard_time_minutes: Optional[float] = None
    frequency: Optional[str] = None
    capacity_unit: Optional[str] = None


class StandardRead(BaseModel):
    id: str
    asset_id: str
    activity_id: str
    product_reference_id: Optional[str] = None
    standard_time_minutes: Optional[float] = None
    frequency: Optional[str] = None
    is_active: bool
    capacity_unit: Optional[str] = None
    asset_name: Optional[str] = None
    activity_name: Optional[str] = None
    reference_code: Optional[str] = None


class ElementCreate(BaseModel):
    name: str
    type: str = "operation"
    is_cyclic: bool = True
    order: int


class StudyCreate(BaseModel):
    name: str
    analyst_name: str
    process_standard_id: Optional[str] = None
    asset_id: Optional[str] = None
    product_reference_id: Optional[str] = None
    study_type: str = "continuous"
    rating_factor: float = 1.0
    supplements_pct: float = 0.0
    confidence_level: float = 0.95
    sampling_interval_seconds: Optional[int] = None
    sampling_population_size: Optional[int] = None
    elements: List[ElementCreate]


class LapRecord(BaseModel):
    element_id: str
    cycle_number: int
    split_time_ms: float
    lap_time_ms: Optional[float] = None
    units_count: int = 1
    is_abnormal: bool = False
    notes: Optional[str] = None


class ApplyStudyToStandardRequest(BaseModel):
    process_standard_id: Optional[str] = None


class BulkImportResponse(BaseModel):
    entity: str
    created: int
    updated: int
    errors_count: int
    errors: List[str]


@router.post("/activities", response_model=ActivityRead)
def create_activity(
    data: ActivityCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    activity = StandardActivity(
        name=data.name.strip(),
        type=data.type.strip(),
        is_value_added=data.is_value_added,
    )
    session.add(activity)
    session.commit()
    session.refresh(activity)
    return ActivityRead(
        id=str(activity.id),
        name=activity.name,
        type=activity.type,
        is_value_added=activity.is_value_added,
    )


@router.get("/activities", response_model=List[ActivityRead])
def list_activities(
    type: Optional[str] = Query(None, description="Filter by activity type"),
    session: Session = Depends(get_session),
):
    stmt = select(StandardActivity)
    if type:
        stmt = stmt.where(StandardActivity.type == type)
    activities = session.exec(stmt).all()
    return [
        ActivityRead(
            id=str(a.id),
            name=a.name,
            type=a.type,
            is_value_added=a.is_value_added,
        )
        for a in activities
    ]


@router.patch("/activities/{activity_id}", response_model=ActivityRead)
def update_activity(
    activity_id: str,
    data: ActivityUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    activity = session.get(StandardActivity, _to_uuid(activity_id, "activity_id"))
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found.")

    if data.name is not None:
        activity.name = data.name.strip()
    if data.type is not None:
        activity.type = data.type.strip()
    if data.is_value_added is not None:
        activity.is_value_added = data.is_value_added

    session.add(activity)
    session.commit()
    session.refresh(activity)
    return ActivityRead(
        id=str(activity.id),
        name=activity.name,
        type=activity.type,
        is_value_added=activity.is_value_added,
    )


@router.delete("/activities/{activity_id}", status_code=204)
def delete_activity(
    activity_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    activity_uuid = _to_uuid(activity_id, "activity_id")
    activity = session.get(StandardActivity, activity_uuid)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found.")

    in_use = session.exec(
        select(ProcessStandard).where(ProcessStandard.activity_id == activity_uuid)
    ).first()
    if in_use:
        raise HTTPException(status_code=409, detail="Activity is used by one or more standards.")

    session.delete(activity)
    session.commit()


@router.post("/references", response_model=ReferenceRead)
def create_reference(
    data: ReferenceCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    existing = session.exec(
        select(ProductReference).where(ProductReference.code == data.code.strip())
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Reference '{data.code}' already exists.")

    ref = ProductReference(
        code=data.code.strip(),
        description=data.description.strip(),
        family=data.family.strip(),
        uom=_text_value(data.uom),
        packaging_uom=_text_value(data.packaging_uom),
    )
    session.add(ref)
    session.commit()
    session.refresh(ref)
    return ReferenceRead(
        id=str(ref.id),
        code=ref.code,
        description=ref.description,
        family=ref.family,
        uom=ref.uom,
        packaging_uom=ref.packaging_uom,
    )


@router.get("/references", response_model=List[ReferenceRead])
def list_references(
    search: Optional[str] = Query(None, description="Search by code or description"),
    family: Optional[str] = Query(None, description="Filter by family"),
    session: Session = Depends(get_session),
):
    stmt = select(ProductReference)
    if family:
        stmt = stmt.where(ProductReference.family == family)
    refs = session.exec(stmt).all()

    results: List[ReferenceRead] = []
    search_q = search.lower().strip() if search else None
    for ref in refs:
        if search_q and search_q not in ref.code.lower() and search_q not in ref.description.lower():
            continue
        results.append(
            ReferenceRead(
                id=str(ref.id),
                code=ref.code,
                description=ref.description,
                family=ref.family,
                uom=ref.uom,
                packaging_uom=ref.packaging_uom,
            )
        )
    return results


@router.patch("/references/{reference_id}", response_model=ReferenceRead)
def update_reference(
    reference_id: str,
    data: ReferenceUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    ref = session.get(ProductReference, _to_uuid(reference_id, "reference_id"))
    if not ref:
        raise HTTPException(status_code=404, detail="Reference not found.")

    if data.code is not None:
        next_code = data.code.strip()
        if next_code != ref.code:
            dup = session.exec(
                select(ProductReference).where(ProductReference.code == next_code)
            ).first()
            if dup:
                raise HTTPException(status_code=409, detail=f"Reference '{next_code}' already exists.")
            ref.code = next_code
    if data.description is not None:
        ref.description = data.description.strip()
    if data.family is not None:
        ref.family = data.family.strip()
    if data.uom is not None:
        ref.uom = _text_value(data.uom)
    if data.packaging_uom is not None:
        ref.packaging_uom = _text_value(data.packaging_uom)

    session.add(ref)
    session.commit()
    session.refresh(ref)
    return ReferenceRead(
        id=str(ref.id),
        code=ref.code,
        description=ref.description,
        family=ref.family,
        uom=ref.uom,
        packaging_uom=ref.packaging_uom,
    )


@router.delete("/references/{reference_id}", status_code=204)
def delete_reference(
    reference_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    ref_uuid = _to_uuid(reference_id, "reference_id")
    ref = session.get(ProductReference, ref_uuid)
    if not ref:
        raise HTTPException(status_code=404, detail="Reference not found.")

    in_standard = session.exec(
        select(ProcessStandard).where(ProcessStandard.product_reference_id == ref_uuid)
    ).first()
    if in_standard:
        raise HTTPException(status_code=409, detail="Reference is used by one or more standards.")

    in_study = session.exec(
        select(TimeStudy).where(TimeStudy.product_reference_id == ref_uuid)
    ).first()
    if in_study:
        raise HTTPException(status_code=409, detail="Reference is used by one or more studies.")

    session.delete(ref)
    session.commit()


@router.post("/standards", response_model=StandardRead)
def create_standard(
    data: StandardCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    asset_uuid = _to_uuid(data.asset_id, "asset_id")
    activity_uuid = _to_uuid(data.activity_id, "activity_id")
    reference_uuid = _to_uuid(data.product_reference_id, "product_reference_id")

    asset = session.get(Asset, asset_uuid)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found.")

    activity = session.get(StandardActivity, activity_uuid)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found.")

    if reference_uuid and not session.get(ProductReference, reference_uuid):
        raise HTTPException(status_code=404, detail="Product reference not found.")

    uniqueness_conditions = [
        ProcessStandard.asset_id == asset_uuid,
        ProcessStandard.activity_id == activity_uuid,
        ProcessStandard.product_reference_id == reference_uuid,
    ]
    existing = session.exec(select(ProcessStandard).where(and_(*uniqueness_conditions))).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail="A standard with this Asset + Activity + Reference already exists.",
        )

    standard = ProcessStandard(
        asset_id=asset_uuid,
        activity_id=activity_uuid,
        product_reference_id=reference_uuid,
        standard_time_minutes=data.standard_time_minutes,
        frequency=_text_value(data.frequency),
        capacity_unit=_text_value(data.capacity_unit),
        is_active=True,
    )
    session.add(standard)
    session.commit()
    session.refresh(standard)
    return _build_standard_read(session, standard)


@router.get("/standards", response_model=List[StandardRead])
def list_standards(
    asset_id: Optional[str] = Query(None, description="Filter by asset UUID"),
    session: Session = Depends(get_session),
):
    stmt = select(ProcessStandard)
    if asset_id:
        stmt = stmt.where(ProcessStandard.asset_id == _to_uuid(asset_id, "asset_id"))
    standards = session.exec(stmt).all()
    return [_build_standard_read(session, standard) for standard in standards]


@router.get("/standards/{standard_id}", response_model=StandardRead)
def get_standard(standard_id: str, session: Session = Depends(get_session)):
    standard = session.get(ProcessStandard, _to_uuid(standard_id, "standard_id"))
    if not standard:
        raise HTTPException(status_code=404, detail="Standard not found.")
    return _build_standard_read(session, standard)


@router.patch("/standards/{standard_id}", response_model=StandardRead)
def update_standard(
    standard_id: str,
    data: StandardUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    standard = session.get(ProcessStandard, _to_uuid(standard_id, "standard_id"))
    if not standard:
        raise HTTPException(status_code=404, detail="Standard not found.")

    if data.is_active is not None:
        standard.is_active = data.is_active
    if data.standard_time_minutes is not None:
        standard.standard_time_minutes = data.standard_time_minutes
    if data.frequency is not None:
        standard.frequency = _text_value(data.frequency)
    if data.capacity_unit is not None:
        standard.capacity_unit = _text_value(data.capacity_unit)

    session.add(standard)
    session.commit()
    session.refresh(standard)
    return _build_standard_read(session, standard)


@router.delete("/standards/{standard_id}", status_code=204)
def delete_standard(
    standard_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    standard_uuid = _to_uuid(standard_id, "standard_id")
    standard = session.get(ProcessStandard, standard_uuid)
    if not standard:
        raise HTTPException(status_code=404, detail="Standard not found.")

    linked_study = session.exec(
        select(TimeStudy).where(TimeStudy.process_standard_id == standard_uuid)
    ).first()
    if linked_study:
        raise HTTPException(status_code=409, detail="Standard is used by one or more studies.")

    session.delete(standard)
    session.commit()


@router.post("/studies", response_model=dict)
def create_time_study(
    data: StudyCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    process_standard_uuid = _to_uuid(data.process_standard_id, "process_standard_id")
    asset_uuid = _to_uuid(data.asset_id, "asset_id")
    reference_uuid = _to_uuid(data.product_reference_id, "product_reference_id")

    standard = None
    if process_standard_uuid:
        standard = session.get(ProcessStandard, process_standard_uuid)
        if not standard:
            raise HTTPException(status_code=404, detail="Process standard not found.")

    effective_asset_uuid = asset_uuid or (standard.asset_id if standard else None)
    effective_reference_uuid = reference_uuid or (standard.product_reference_id if standard else None)

    if standard and asset_uuid and standard.asset_id != asset_uuid:
        raise HTTPException(
            status_code=400,
            detail="asset_id does not match the selected process standard.",
        )
    if standard and reference_uuid and standard.product_reference_id != reference_uuid:
        raise HTTPException(
            status_code=400,
            detail="product_reference_id does not match the selected process standard.",
        )

    if effective_asset_uuid and not session.get(Asset, effective_asset_uuid):
        raise HTTPException(status_code=404, detail="Asset not found.")
    if effective_reference_uuid and not session.get(ProductReference, effective_reference_uuid):
        raise HTTPException(status_code=404, detail="Product reference not found.")

    study = TimeStudy(
        name=data.name.strip(),
        analyst_name=data.analyst_name.strip(),
        study_type=data.study_type,
        rating_factor=data.rating_factor,
        supplements_pct=data.supplements_pct,
        confidence_level=data.confidence_level,
        sampling_interval_seconds=data.sampling_interval_seconds,
        sampling_population_size=data.sampling_population_size,
        status="draft",
        process_standard_id=process_standard_uuid,
        asset_id=effective_asset_uuid,
        product_reference_id=effective_reference_uuid,
    )
    session.add(study)
    session.commit()
    session.refresh(study)

    for element_data in sorted(data.elements, key=lambda element: element.order):
        session.add(
            TimingElement(
                time_study_id=study.id,
                name=element_data.name.strip(),
                type=element_data.type,
                is_cyclic=element_data.is_cyclic,
                order=element_data.order,
            )
        )
    session.commit()

    return {
        "id": str(study.id),
        "name": study.name,
        "status": study.status,
        "elements_count": len(data.elements),
        "process_standard_id": str(study.process_standard_id) if study.process_standard_id else None,
        "asset_id": str(study.asset_id) if study.asset_id else None,
        "product_reference_id": str(study.product_reference_id) if study.product_reference_id else None,
    }


@router.get("/studies", response_model=List[dict])
def list_time_studies(session: Session = Depends(get_session)):
    studies = session.exec(select(TimeStudy).order_by(TimeStudy.study_date.desc())).all()  # type: ignore
    result = []
    for study in studies:
        standard = (
            session.get(ProcessStandard, study.process_standard_id)
            if study.process_standard_id
            else None
        )
        linked_asset_id = _study_linked_asset_id(study, standard)
        linked_reference_id = _study_linked_reference_id(study, standard)
        linked_asset = session.get(Asset, linked_asset_id) if linked_asset_id else None
        linked_reference = (
            session.get(ProductReference, linked_reference_id) if linked_reference_id else None
        )
        result.append(
            {
                "id": str(study.id),
                "name": study.name,
                "analyst_name": study.analyst_name,
                "study_type": study.study_type,
                "status": study.status,
                "study_date": study.study_date.isoformat(),
                "confidence_level": study.confidence_level,
                "sampling_interval_seconds": study.sampling_interval_seconds,
                "sampling_population_size": study.sampling_population_size,
                "process_standard_id": str(study.process_standard_id)
                if study.process_standard_id
                else None,
                "asset_id": str(linked_asset_id) if linked_asset_id else None,
                "asset_name": linked_asset.name if linked_asset else None,
                "product_reference_id": str(linked_reference_id) if linked_reference_id else None,
                "reference_code": linked_reference.code if linked_reference else None,
            }
        )
    return result


@router.get("/studies/{study_id}", response_model=dict)
def get_time_study(study_id: str, session: Session = Depends(get_session)):
    study = session.get(TimeStudy, _to_uuid(study_id, "study_id"))
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    elements = session.exec(
        select(TimingElement)
        .where(TimingElement.time_study_id == study.id)
        .order_by(TimingElement.order)
    ).all()

    sessions_db = session.exec(
        select(TimingSession).where(TimingSession.time_study_id == study.id)
    ).all()

    sessions_data = []
    for timing_session in sessions_db:
        laps = session.exec(select(TimingLap).where(TimingLap.session_id == timing_session.id)).all()
        sessions_data.append(
            {
                "id": str(timing_session.id),
                "started_at": timing_session.started_at.isoformat(),
                "ended_at": timing_session.ended_at.isoformat() if timing_session.ended_at else None,
                "laps_count": len(laps),
                "laps": [
                    {
                        "id": str(lap.id),
                        "element_id": str(lap.element_id),
                        "cycle_number": lap.cycle_number,
                        "split_time_ms": lap.split_time_ms,
                        "lap_time_ms": lap.lap_time_ms,
                        "units_count": lap.units_count,
                        "is_abnormal": lap.is_abnormal,
                        "notes": lap.notes,
                    }
                    for lap in laps
                ],
            }
        )

    standard = session.get(ProcessStandard, study.process_standard_id) if study.process_standard_id else None
    linked_asset_id = _study_linked_asset_id(study, standard)
    linked_reference_id = _study_linked_reference_id(study, standard)
    linked_asset = session.get(Asset, linked_asset_id) if linked_asset_id else None
    linked_reference = session.get(ProductReference, linked_reference_id) if linked_reference_id else None

    return {
        "id": str(study.id),
        "name": study.name,
        "analyst_name": study.analyst_name,
        "study_type": study.study_type,
        "status": study.status,
        "rating_factor": study.rating_factor,
        "supplements_pct": study.supplements_pct,
        "confidence_level": study.confidence_level,
        "sampling_interval_seconds": study.sampling_interval_seconds,
        "sampling_population_size": study.sampling_population_size,
        "process_standard_id": str(study.process_standard_id) if study.process_standard_id else None,
        "asset_id": str(linked_asset_id) if linked_asset_id else None,
        "asset_name": linked_asset.name if linked_asset else None,
        "product_reference_id": str(linked_reference_id) if linked_reference_id else None,
        "reference_code": linked_reference.code if linked_reference else None,
        "elements": [
            {
                "id": str(element.id),
                "name": element.name,
                "type": element.type,
                "is_cyclic": element.is_cyclic,
                "order": element.order,
            }
            for element in elements
        ],
        "sessions": sessions_data,
    }


@router.delete("/studies/{study_id}", status_code=204)
def delete_time_study(
    study_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    study = session.get(TimeStudy, _to_uuid(study_id, "study_id"))
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    sessions_db = session.exec(
        select(TimingSession).where(TimingSession.time_study_id == study.id)
    ).all()
    for timing_session in sessions_db:
        laps = session.exec(
            select(TimingLap).where(TimingLap.session_id == timing_session.id)
        ).all()
        for lap in laps:
            session.delete(lap)
        session.delete(timing_session)

    elements = session.exec(
        select(TimingElement).where(TimingElement.time_study_id == study.id)
    ).all()
    for element in elements:
        session.delete(element)

    session.delete(study)
    session.commit()


@router.post("/studies/{study_id}/sessions", response_model=dict)
def start_session(
    study_id: str,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    study = session.get(TimeStudy, _to_uuid(study_id, "study_id"))
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    study.status = "in_progress"
    timing_session = TimingSession(time_study_id=study.id)
    session.add(study)
    session.add(timing_session)
    session.commit()
    session.refresh(timing_session)

    return {
        "session_id": str(timing_session.id),
        "started_at": timing_session.started_at.isoformat(),
    }


@router.post("/studies/{study_id}/laps", response_model=dict)
def record_lap(
    study_id: str,
    data: LapRecord,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    study = session.get(TimeStudy, _to_uuid(study_id, "study_id"))
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    element_uuid = _to_uuid(data.element_id, "element_id")
    element = session.get(TimingElement, element_uuid)
    if not element or element.time_study_id != study.id:
        raise HTTPException(status_code=400, detail="Element does not belong to this study.")

    active_sessions = session.exec(
        select(TimingSession)
        .where(TimingSession.time_study_id == study.id)
        .where(TimingSession.ended_at == None)  # noqa: E711
        .order_by(TimingSession.started_at.desc())  # type: ignore
    ).all()
    if not active_sessions:
        raise HTTPException(status_code=400, detail="No active session. Start one first.")

    active_session = active_sessions[0]
    lap = TimingLap(
        session_id=active_session.id,
        element_id=element_uuid,
        cycle_number=data.cycle_number,
        split_time_ms=data.split_time_ms,
        lap_time_ms=data.lap_time_ms,
        units_count=data.units_count,
        is_abnormal=data.is_abnormal,
        notes=data.notes,
    )
    session.add(lap)
    session.commit()
    session.refresh(lap)

    return {
        "lap_id": str(lap.id),
        "cycle_number": lap.cycle_number,
        "split_time_ms": lap.split_time_ms,
    }


@router.get("/studies/{study_id}/results", response_model=dict)
def get_study_results(study_id: str, session: Session = Depends(get_session)):
    from ..services.timing_engine import TimingEngine

    engine = TimingEngine(session)
    result = engine.calculate_results(_to_uuid(study_id, "study_id"))
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


@router.post("/studies/{study_id}/apply-to-standard", response_model=dict)
def apply_results_to_standard(
    study_id: str,
    payload: ApplyStudyToStandardRequest,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    study_uuid = _to_uuid(study_id, "study_id")
    study = session.get(TimeStudy, study_uuid)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    from ..services.timing_engine import TimingEngine

    engine = TimingEngine(session)
    results = engine.calculate_results(study_uuid)
    if "error" in results:
        raise HTTPException(status_code=400, detail=results["error"])

    target_standard_uuid = _to_uuid(payload.process_standard_id, "process_standard_id")
    if not target_standard_uuid:
        target_standard_uuid = study.process_standard_id
    if not target_standard_uuid:
        raise HTTPException(
            status_code=400,
            detail="Study is not linked to a process standard. Provide process_standard_id.",
        )

    standard = session.get(ProcessStandard, target_standard_uuid)
    if not standard:
        raise HTTPException(status_code=404, detail="Process standard not found")

    standard.standard_time_minutes = results["total_standard_time_minutes"]
    study.process_standard_id = target_standard_uuid
    study.status = "completed"
    study.calculated_normal_time = results["total_normal_time_ms"] / 60000
    study.calculated_standard_time = results["total_standard_time_minutes"]

    session.add(standard)
    session.add(study)
    session.commit()
    session.refresh(standard)
    session.refresh(study)

    return {
        "message": "Study results applied to standard successfully.",
        "study_id": str(study.id),
        "process_standard_id": str(standard.id),
        "standard_time_minutes": standard.standard_time_minutes,
        "study_status": study.status,
    }


def _build_template_workbook(entity: str, session: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active

    if entity == "references":
        ws.title = "references"
        ws.append(["code", "description", "family", "uom", "packaging_uom"])
        ws.append(["SKU-001", "Producto ejemplo", "Familia A", "kg", "caja"])
    elif entity == "activities":
        ws.title = "activities"
        ws.append(["name", "type", "is_value_added"])
        ws.append(["Ensamble", "Operation", True])
    elif entity == "standards":
        ws.title = "standards"
        ws.append(
            [
                "asset_id",
                "activity_id",
                "product_reference_id",
                "standard_time_minutes",
                "frequency",
                "is_active",
                "capacity_unit",
            ]
        )
        ws.append(
            [
                "uuid_asset",
                "uuid_activity",
                "uuid_reference_or_empty",
                1.5,
                "Per Unit",
                True,
                "und/h",
            ]
        )
    else:
        raise HTTPException(status_code=400, detail="Unsupported entity for template.")

    catalog = wb.create_sheet("catalog")
    catalog.append(["type", "id", "name", "extra"])
    for asset in session.exec(select(Asset)).all():
        catalog.append(["asset", str(asset.id), asset.name, asset.type])
    for activity in session.exec(select(StandardActivity)).all():
        catalog.append(["activity", str(activity.id), activity.name, activity.type])
    for reference in session.exec(select(ProductReference)).all():
        catalog.append(["reference", str(reference.id), reference.code, reference.description])
    return wb


def _build_export_workbook(entity: str, session: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active

    if entity == "references":
        ws.title = "references"
        ws.append(["id", "code", "description", "family", "uom", "packaging_uom"])
        for ref in session.exec(select(ProductReference)).all():
            ws.append(
                [
                    str(ref.id),
                    ref.code,
                    ref.description,
                    ref.family,
                    ref.uom,
                    ref.packaging_uom,
                ]
            )
        return wb

    if entity == "activities":
        ws.title = "activities"
        ws.append(["id", "name", "type", "is_value_added"])
        for activity in session.exec(select(StandardActivity)).all():
            ws.append([str(activity.id), activity.name, activity.type, activity.is_value_added])
        return wb

    if entity == "standards":
        ws.title = "standards"
        ws.append(
            [
                "id",
                "asset_id",
                "activity_id",
                "product_reference_id",
                "standard_time_minutes",
                "frequency",
                "is_active",
                "capacity_unit",
            ]
        )
        for standard in session.exec(select(ProcessStandard)).all():
            ws.append(
                [
                    str(standard.id),
                    str(standard.asset_id),
                    str(standard.activity_id),
                    str(standard.product_reference_id) if standard.product_reference_id else None,
                    standard.standard_time_minutes,
                    standard.frequency,
                    standard.is_active,
                    standard.capacity_unit,
                ]
            )
        return wb

    raise HTTPException(status_code=400, detail="Unsupported entity for export.")


def _import_references(rows: List[tuple], session: Session) -> tuple[int, int, List[str]]:
    created = 0
    updated = 0
    errors: List[str] = []
    for row_idx, row in enumerate(rows, start=2):
        code = _text_value(row[0] if len(row) > 0 else None)
        description = _text_value(row[1] if len(row) > 1 else None)
        family = _text_value(row[2] if len(row) > 2 else None)
        uom = _text_value(row[3] if len(row) > 3 else None)
        packaging_uom = _text_value(row[4] if len(row) > 4 else None)

        if not code and not description and not family:
            continue
        if not code or not description or not family:
            errors.append(f"Row {row_idx}: code, description and family are required.")
            continue

        existing = session.exec(
            select(ProductReference).where(ProductReference.code == code)
        ).first()
        if existing:
            existing.description = description
            existing.family = family
            existing.uom = uom
            existing.packaging_uom = packaging_uom
            session.add(existing)
            updated += 1
        else:
            session.add(
                ProductReference(
                    code=code,
                    description=description,
                    family=family,
                    uom=uom,
                    packaging_uom=packaging_uom,
                )
            )
            created += 1
    return created, updated, errors


def _import_activities(rows: List[tuple], session: Session) -> tuple[int, int, List[str]]:
    created = 0
    updated = 0
    errors: List[str] = []
    for row_idx, row in enumerate(rows, start=2):
        name = _text_value(row[0] if len(row) > 0 else None)
        activity_type = _text_value(row[1] if len(row) > 1 else None)
        is_value_added = _bool_value(row[2] if len(row) > 2 else None, default=False)

        if not name and not activity_type:
            continue
        if not name or not activity_type:
            errors.append(f"Row {row_idx}: name and type are required.")
            continue

        existing = session.exec(
            select(StandardActivity).where(
                and_(StandardActivity.name == name, StandardActivity.type == activity_type)
            )
        ).first()
        if existing:
            existing.is_value_added = is_value_added
            session.add(existing)
            updated += 1
        else:
            session.add(
                StandardActivity(
                    name=name,
                    type=activity_type,
                    is_value_added=is_value_added,
                )
            )
            created += 1
    return created, updated, errors


def _import_standards(rows: List[tuple], session: Session) -> tuple[int, int, List[str]]:
    created = 0
    updated = 0
    errors: List[str] = []
    for row_idx, row in enumerate(rows, start=2):
        asset_raw = _text_value(row[0] if len(row) > 0 else None)
        activity_raw = _text_value(row[1] if len(row) > 1 else None)
        reference_raw = _text_value(row[2] if len(row) > 2 else None)
        standard_time = _float_value(row[3] if len(row) > 3 else None)
        frequency = _text_value(row[4] if len(row) > 4 else None)
        is_active = _bool_value(row[5] if len(row) > 5 else True, default=True)
        capacity_unit = _text_value(row[6] if len(row) > 6 else None)

        if not asset_raw and not activity_raw:
            continue
        if not asset_raw or not activity_raw:
            errors.append(f"Row {row_idx}: asset_id and activity_id are required.")
            continue

        try:
            asset_uuid = uuid_lib.UUID(asset_raw)
            activity_uuid = uuid_lib.UUID(activity_raw)
            reference_uuid = uuid_lib.UUID(reference_raw) if reference_raw else None
        except ValueError:
            errors.append(f"Row {row_idx}: invalid UUID.")
            continue

        if not session.get(Asset, asset_uuid):
            errors.append(f"Row {row_idx}: asset '{asset_uuid}' does not exist.")
            continue
        if not session.get(StandardActivity, activity_uuid):
            errors.append(f"Row {row_idx}: activity '{activity_uuid}' does not exist.")
            continue
        if reference_uuid and not session.get(ProductReference, reference_uuid):
            errors.append(f"Row {row_idx}: reference '{reference_uuid}' does not exist.")
            continue

        existing = session.exec(
            select(ProcessStandard).where(
                and_(
                    ProcessStandard.asset_id == asset_uuid,
                    ProcessStandard.activity_id == activity_uuid,
                    ProcessStandard.product_reference_id == reference_uuid,
                )
            )
        ).first()
        if existing:
            existing.standard_time_minutes = standard_time
            existing.frequency = frequency
            existing.is_active = is_active
            existing.capacity_unit = capacity_unit
            session.add(existing)
            updated += 1
        else:
            session.add(
                ProcessStandard(
                    asset_id=asset_uuid,
                    activity_id=activity_uuid,
                    product_reference_id=reference_uuid,
                    standard_time_minutes=standard_time,
                    frequency=frequency,
                    is_active=is_active,
                    capacity_unit=capacity_unit,
                )
            )
            created += 1
    return created, updated, errors


@router.get("/xlsx/template")
def download_xlsx_template(
    entity: str = Query(..., pattern="^(references|activities|standards)$"),
    session: Session = Depends(get_session),
):
    workbook = _build_template_workbook(entity, session)
    return _workbook_response(workbook, f"takta_{entity}_template.xlsx")


@router.get("/xlsx/export")
def download_xlsx_export(
    entity: str = Query(..., pattern="^(references|activities|standards)$"),
    session: Session = Depends(get_session),
):
    workbook = _build_export_workbook(entity, session)
    return _workbook_response(workbook, f"takta_{entity}_export.xlsx")


@router.post("/xlsx/import", response_model=BulkImportResponse)
def import_xlsx_data(
    entity: str = Query(..., pattern="^(references|activities|standards)$"),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file.")

    try:
        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLSX file.") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    if entity == "references":
        created, updated, errors = _import_references(rows, session)
    elif entity == "activities":
        created, updated, errors = _import_activities(rows, session)
    else:
        created, updated, errors = _import_standards(rows, session)

    session.commit()
    return BulkImportResponse(
        entity=entity,
        created=created,
        updated=updated,
        errors_count=len(errors),
        errors=errors,
    )
