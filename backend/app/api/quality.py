from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from statistics import mean, pstdev
from typing import Any, Dict, List, Optional
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field as PydanticField
from sqlmodel import Session, select

from ..core.auth import CurrentUser, get_current_user, require_role
from ..db import get_session
from ..models import (
    Asset,
    CapaAction,
    ImprovementAction,
    NonConformity,
    ProcessStandard,
    ProductReference,
    StandardActivity,
    WeightCapabilityRun,
    WeightSample,
    WeightSamplingSpec,
)
from ..services.reference_guard import validate_canonical_context


router = APIRouter(
    prefix="/api/quality",
    tags=["Quality"],
    dependencies=[Depends(get_current_user)],
)


D2_FOR_MOVING_RANGE = 1.128


def _to_uuid(raw: Optional[str], field_name: str) -> Optional[uuid.UUID]:
    if raw in (None, ""):
        return None
    try:
        return uuid.UUID(str(raw))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=f"Invalid UUID for '{field_name}'.") from exc


def _clean_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _as_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "si"}


def _parse_datetime_cell(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return datetime.utcnow()
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return datetime.utcnow()


def _validate_limits(
    lower_limit: float,
    upper_limit: float,
    target_weight: Optional[float],
) -> tuple[float, float]:
    if upper_limit <= lower_limit:
        raise HTTPException(status_code=400, detail="upper_limit must be greater than lower_limit.")
    if target_weight is not None and not (lower_limit <= target_weight <= upper_limit):
        raise HTTPException(status_code=400, detail="target_weight must be inside lower_limit and upper_limit.")
    return lower_limit, upper_limit


def _resolve_links(
    session: Session,
    asset_id: Optional[uuid.UUID],
    product_reference_id: Optional[uuid.UUID],
    process_standard_id: Optional[uuid.UUID],
) -> tuple[Optional[uuid.UUID], Optional[uuid.UUID], Optional[uuid.UUID]]:
    context = validate_canonical_context(
        session=session,
        asset_id=asset_id,
        product_reference_id=product_reference_id,
        process_standard_id=process_standard_id,
    )
    return context.asset_id, context.product_reference_id, context.process_standard_id


def _target_for_spec(spec: WeightSamplingSpec) -> float:
    if spec.target_weight is not None:
        return spec.target_weight
    return (spec.lower_limit + spec.upper_limit) / 2


def _warning_band(spec: WeightSamplingSpec) -> float:
    spread = max(spec.upper_limit - spec.lower_limit, 0.0)
    pct = min(max(spec.warning_band_pct, 0.0), 0.49)
    return spread * pct


def _status_for_value(spec: WeightSamplingSpec, measured_value: float) -> str:
    if measured_value < spec.lower_limit or measured_value > spec.upper_limit:
        return "red"
    band = _warning_band(spec)
    if band > 0 and (measured_value <= spec.lower_limit + band or measured_value >= spec.upper_limit - band):
        return "yellow"
    return "green"


def _sample_payload(row: WeightSample) -> Dict[str, Any]:
    return {
        "id": str(row.id),
        "specification_id": str(row.specification_id),
        "measured_value": row.measured_value,
        "measured_at": row.measured_at.isoformat(),
        "measured_by": row.measured_by,
        "batch_code": row.batch_code,
        "shift": row.shift,
        "notes": row.notes,
        "deviation": row.deviation,
        "status_color": row.status_color,
    }


def _build_summary(spec: WeightSamplingSpec, samples: List[WeightSample]) -> Dict[str, Any]:
    values = [row.measured_value for row in samples]
    total = len(values)
    in_spec = sum(1 for value in values if spec.lower_limit <= value <= spec.upper_limit)
    status_breakdown = {
        "green": sum(1 for row in samples if row.status_color == "green"),
        "yellow": sum(1 for row in samples if row.status_color == "yellow"),
        "red": sum(1 for row in samples if row.status_color == "red"),
    }

    return {
        "samples_count": total,
        "in_spec_count": in_spec,
        "in_spec_pct": round((in_spec / total) * 100, 2) if total else 0.0,
        "avg": round(mean(values), 6) if total else None,
        "min": min(values) if total else None,
        "max": max(values) if total else None,
        "stddev": round(pstdev(values), 6) if total > 1 else 0.0,
        "status_breakdown": status_breakdown,
    }


def _moving_ranges(values: List[float]) -> List[float]:
    if len(values) < 2:
        return []
    return [abs(values[idx] - values[idx - 1]) for idx in range(1, len(values))]


def _compute_spc_stats(values: List[float]) -> Dict[str, Optional[float]]:
    if not values:
        return {
            "center_line": None,
            "sigma_within": None,
            "sigma_overall": None,
            "ucl": None,
            "lcl": None,
            "mr_bar": None,
        }

    center_line = float(mean(values))
    moving_ranges = _moving_ranges(values)
    mr_bar = float(mean(moving_ranges)) if moving_ranges else 0.0
    sigma_within = (mr_bar / D2_FOR_MOVING_RANGE) if mr_bar > 0 else 0.0
    sigma_overall = float(pstdev(values)) if len(values) > 1 else 0.0

    if sigma_within <= 0 and sigma_overall > 0:
        sigma_within = sigma_overall

    if sigma_within > 0:
        ucl = center_line + (3 * sigma_within)
        lcl = center_line - (3 * sigma_within)
    else:
        ucl = center_line
        lcl = center_line

    return {
        "center_line": round(center_line, 6),
        "sigma_within": round(sigma_within, 6),
        "sigma_overall": round(sigma_overall, 6),
        "ucl": round(ucl, 6),
        "lcl": round(lcl, 6),
        "mr_bar": round(mr_bar, 6),
    }


def _detect_western_electric_rules(
    values: List[float],
    sample_ids: List[str],
    center_line: Optional[float],
    sigma: Optional[float],
) -> List[Dict[str, Any]]:
    if not values or center_line is None or sigma is None or sigma <= 0:
        return []

    violations: List[Dict[str, Any]] = []
    upper_1 = center_line + sigma
    upper_2 = center_line + (2 * sigma)
    upper_3 = center_line + (3 * sigma)
    lower_1 = center_line - sigma
    lower_2 = center_line - (2 * sigma)
    lower_3 = center_line - (3 * sigma)

    # WE Rule 1: one point beyond 3 sigma.
    for idx, value in enumerate(values):
        if value > upper_3 or value < lower_3:
            violations.append(
                {
                    "rule_code": "WE1",
                    "rule_name": "1 punto fuera de 3 sigma",
                    "severity": "critical",
                    "sample_ids": [sample_ids[idx]],
                    "indices": [idx],
                    "message": "Punto fuera de limites de control de 3 sigma.",
                }
            )

    # WE Rule 2: two of three consecutive beyond 2 sigma on same side.
    for start in range(0, len(values) - 2):
        window = values[start : start + 3]
        idxs = list(range(start, start + 3))
        above = [idx for idx, value in zip(idxs, window) if value > upper_2]
        below = [idx for idx, value in zip(idxs, window) if value < lower_2]
        selected = above if len(above) >= 2 else below if len(below) >= 2 else []
        if selected:
            violations.append(
                {
                    "rule_code": "WE2",
                    "rule_name": "2 de 3 puntos sobre 2 sigma",
                    "severity": "warning",
                    "sample_ids": [sample_ids[idx] for idx in selected],
                    "indices": selected,
                    "message": "Dos de tres puntos consecutivos exceden 2 sigma del mismo lado.",
                }
            )

    # WE Rule 3: four of five consecutive beyond 1 sigma on same side.
    for start in range(0, len(values) - 4):
        window = values[start : start + 5]
        idxs = list(range(start, start + 5))
        above = [idx for idx, value in zip(idxs, window) if value > upper_1]
        below = [idx for idx, value in zip(idxs, window) if value < lower_1]
        selected = above if len(above) >= 4 else below if len(below) >= 4 else []
        if selected:
            violations.append(
                {
                    "rule_code": "WE3",
                    "rule_name": "4 de 5 puntos sobre 1 sigma",
                    "severity": "warning",
                    "sample_ids": [sample_ids[idx] for idx in selected],
                    "indices": selected,
                    "message": "Cuatro de cinco puntos consecutivos exceden 1 sigma del mismo lado.",
                }
            )

    # WE Rule 4: eight consecutive points on one side of centerline.
    for start in range(0, len(values) - 7):
        window = values[start : start + 8]
        idxs = list(range(start, start + 8))
        all_above = all(value > center_line for value in window)
        all_below = all(value < center_line for value in window)
        if all_above or all_below:
            violations.append(
                {
                    "rule_code": "WE4",
                    "rule_name": "8 puntos del mismo lado de la media",
                    "severity": "warning",
                    "sample_ids": [sample_ids[idx] for idx in idxs],
                    "indices": idxs,
                    "message": "Ocho puntos consecutivos se ubican al mismo lado de la linea central.",
                }
            )

    return violations


def _spc_alert_level(violations: List[Dict[str, Any]]) -> str:
    if not violations:
        return "healthy"
    if any(row.get("severity") == "critical" for row in violations):
        return "critical"
    return "warning"


def _build_spc_points(
    samples: List[WeightSample],
    center_line: Optional[float],
    sigma_within: Optional[float],
) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    prev_value: Optional[float] = None
    for idx, sample in enumerate(samples):
        mr = abs(sample.measured_value - prev_value) if prev_value is not None else None
        if sigma_within and sigma_within > 0 and center_line is not None:
            z_score = (sample.measured_value - center_line) / sigma_within
            beyond_3sigma = abs(z_score) > 3
        else:
            z_score = None
            beyond_3sigma = False

        points.append(
            {
                "index": idx,
                "sample_id": str(sample.id),
                "measured_value": sample.measured_value,
                "measured_at": sample.measured_at.isoformat(),
                "measured_by": sample.measured_by,
                "batch_code": sample.batch_code,
                "shift": sample.shift,
                "notes": sample.notes,
                "status_color": sample.status_color,
                "deviation": sample.deviation,
                "moving_range": round(mr, 6) if mr is not None else None,
                "z_score": round(z_score, 6) if z_score is not None else None,
                "beyond_3sigma": beyond_3sigma,
            }
        )
        prev_value = sample.measured_value
    return points


def _build_capability_payload(
    spec: WeightSamplingSpec,
    values: List[float],
    sigma_within: Optional[float],
    sigma_overall: Optional[float],
) -> Dict[str, Any]:
    if not values:
        return {
            "cp": None,
            "cpk": None,
            "pp": None,
            "ppk": None,
            "process_mean": None,
            "centered_mean": None,
            "status": "insufficient_data",
            "message": "No hay muestras para calcular capacidad.",
        }

    process_mean = float(mean(values))
    spread = float(spec.upper_limit - spec.lower_limit)

    def _safe_cp(denominator: Optional[float]) -> Optional[float]:
        if not denominator or denominator <= 0:
            return None
        return round(spread / (6 * denominator), 6)

    def _safe_cpk(denominator: Optional[float]) -> Optional[float]:
        if not denominator or denominator <= 0:
            return None
        upper = (spec.upper_limit - process_mean) / (3 * denominator)
        lower = (process_mean - spec.lower_limit) / (3 * denominator)
        return round(min(upper, lower), 6)

    cp = _safe_cp(sigma_within)
    cpk = _safe_cpk(sigma_within)
    pp = _safe_cp(sigma_overall)
    ppk = _safe_cpk(sigma_overall)

    capability_index = cpk if cpk is not None else ppk
    if capability_index is None:
        status = "insufficient_data"
        message = "No hay variacion suficiente para calcular capacidad."
    elif capability_index >= 1.33:
        status = "capable"
        message = "Proceso capaz (indice >= 1.33)."
    elif capability_index >= 1.0:
        status = "marginal"
        message = "Proceso marginal (indice entre 1.00 y 1.33)."
    else:
        status = "not_capable"
        message = "Proceso no capaz (indice < 1.00)."

    target = _target_for_spec(spec)
    centered_mean = round(process_mean - target, 6)
    return {
        "cp": cp,
        "cpk": cpk,
        "pp": pp,
        "ppk": ppk,
        "process_mean": round(process_mean, 6),
        "centered_mean": centered_mean,
        "status": status,
        "message": message,
    }


def _spec_payload(session: Session, row: WeightSamplingSpec, include_summary: bool = False) -> Dict[str, Any]:
    asset = session.get(Asset, row.asset_id) if row.asset_id else None
    reference = session.get(ProductReference, row.product_reference_id) if row.product_reference_id else None
    standard = session.get(ProcessStandard, row.process_standard_id) if row.process_standard_id else None
    activity = session.get(StandardActivity, standard.activity_id) if standard else None

    last_sample = session.exec(
        select(WeightSample)
        .where(WeightSample.specification_id == row.id)
        .order_by(WeightSample.measured_at.desc())  # type: ignore
        .limit(1)
    ).first()
    samples_count = session.exec(
        select(WeightSample).where(WeightSample.specification_id == row.id)
    ).all()

    payload: Dict[str, Any] = {
        "id": str(row.id),
        "name": row.name,
        "asset_id": str(row.asset_id) if row.asset_id else None,
        "asset_name": asset.name if asset else None,
        "product_reference_id": str(row.product_reference_id) if row.product_reference_id else None,
        "reference_code": reference.code if reference else None,
        "process_standard_id": str(row.process_standard_id) if row.process_standard_id else None,
        "standard_label": activity.name if activity else None,
        "unit": row.unit,
        "lower_limit": row.lower_limit,
        "target_weight": row.target_weight,
        "upper_limit": row.upper_limit,
        "warning_band_pct": row.warning_band_pct,
        "sample_size": row.sample_size,
        "notes": row.notes,
        "is_active": row.is_active,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat(),
        "updated_at": row.updated_at.isoformat(),
        "samples_count": len(samples_count),
        "last_status_color": last_sample.status_color if last_sample else None,
        "last_measured_at": last_sample.measured_at.isoformat() if last_sample else None,
    }

    if include_summary:
        payload["summary"] = _build_summary(spec=row, samples=samples_count)
    return payload


def _workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


class WeightSpecCreate(BaseModel):
    name: str
    asset_id: Optional[uuid.UUID] = None
    product_reference_id: Optional[uuid.UUID] = None
    process_standard_id: Optional[uuid.UUID] = None
    unit: str = "g"
    lower_limit: float
    target_weight: Optional[float] = None
    upper_limit: float
    warning_band_pct: float = PydanticField(default=0.1, ge=0.0, le=0.49)
    sample_size: int = PydanticField(default=5, ge=1, le=5000)
    notes: Optional[str] = None
    is_active: bool = True


class WeightSpecUpdate(BaseModel):
    name: Optional[str] = None
    asset_id: Optional[uuid.UUID] = None
    product_reference_id: Optional[uuid.UUID] = None
    process_standard_id: Optional[uuid.UUID] = None
    unit: Optional[str] = None
    lower_limit: Optional[float] = None
    target_weight: Optional[float] = None
    upper_limit: Optional[float] = None
    warning_band_pct: Optional[float] = PydanticField(default=None, ge=0.0, le=0.49)
    sample_size: Optional[int] = PydanticField(default=None, ge=1, le=5000)
    notes: Optional[str] = None
    is_active: Optional[bool] = None


class WeightSampleCreate(BaseModel):
    measured_value: float
    measured_at: Optional[datetime] = None
    measured_by: Optional[str] = None
    batch_code: Optional[str] = None
    shift: Optional[str] = None
    notes: Optional[str] = None
    auto_create_non_conformity: bool = True
    minimum_alert_level: str = "warning"


class WeightSampleUpdate(BaseModel):
    measured_value: Optional[float] = None
    measured_at: Optional[datetime] = None
    measured_by: Optional[str] = None
    batch_code: Optional[str] = None
    shift: Optional[str] = None
    notes: Optional[str] = None
    auto_create_non_conformity: bool = True
    minimum_alert_level: str = "warning"


class NonConformityCreate(BaseModel):
    asset_id: Optional[uuid.UUID] = None
    product_reference_id: Optional[uuid.UUID] = None
    process_standard_id: Optional[uuid.UUID] = None
    weight_specification_id: Optional[uuid.UUID] = None
    weight_sample_id: Optional[uuid.UUID] = None
    source: str = "manual"
    severity: str = "medium"
    status: str = "Open"
    title: str
    description: Optional[str] = None
    root_cause: Optional[str] = None
    containment: Optional[str] = None


class NonConformityUpdate(BaseModel):
    severity: Optional[str] = None
    status: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    root_cause: Optional[str] = None
    containment: Optional[str] = None
    rejected_reason: Optional[str] = None


class CapaActionCreate(BaseModel):
    action_type: str = "Corrective"
    title: str
    description: Optional[str] = None
    responsible: str
    due_date: Optional[date] = None
    status: str = "Open"
    auto_link_improvement_action: bool = True


class CapaActionUpdate(BaseModel):
    action_type: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    responsible: Optional[str] = None
    due_date: Optional[date] = None
    status: Optional[str] = None
    verification_notes: Optional[str] = None
    rejected_reason: Optional[str] = None


class CapabilityRunCreate(BaseModel):
    limit: int = PydanticField(default=300, ge=20, le=5000)
    run_type: str = "on_demand"
    auto_link_improvement_action: bool = True
    notes: Optional[str] = None


class CapabilityBatchCreate(BaseModel):
    limit: int = PydanticField(default=300, ge=20, le=5000)
    asset_id: Optional[uuid.UUID] = None
    run_type: str = "batch"
    only_active: bool = True
    auto_link_improvement_action: bool = True
    notes: Optional[str] = None


class NonConformityAutoGenerateRequest(BaseModel):
    limit: int = PydanticField(default=300, ge=20, le=2000)
    only_active_specs: bool = True
    minimum_alert_level: str = "warning"  # warning | critical


def _normalize_severity(value: Optional[str], default: str = "medium") -> str:
    raw = (value or default).strip().lower()
    if raw in {"low", "medium", "high", "critical"}:
        return raw
    return default


def _normalize_workflow_status(value: Optional[str], default: str = "Open") -> str:
    raw = (value or default).strip().lower()
    if raw in {"open", "abierta", "abierto"}:
        return "Open"
    if raw in {"in progress", "in_progress", "en progreso", "progreso"}:
        return "In Progress"
    if raw in {"close requested", "close_requested", "solicita cierre", "solicitud cierre", "closerequested"}:
        return "Close Requested"
    if raw in {"approved", "aprobado", "aprobada"}:
        return "Approved"
    if raw in {"rejected", "rechazado", "rechazada"}:
        return "Rejected"
    if raw in {"closed", "cerrada", "cerrado"}:
        return "Closed"
    if raw in {"verified", "verificada", "verificado"}:
        return "Verified"
    return default


def _is_closed_status(status: str) -> bool:
    return status in {"Approved", "Closed", "Verified"}


def _open_workflow_statuses() -> set[str]:
    return {"Open", "In Progress", "Close Requested", "Approved", "Rejected"}


def _normalize_alert_level(value: Optional[str]) -> str:
    raw = (value or "warning").strip().lower()
    return "critical" if raw == "critical" else "warning"


def _minimum_alert_rank(level: str) -> int:
    if level == "critical":
        return 2
    if level == "warning":
        return 1
    return 0


def _workflow_transition_allowed(
    entity: str,
    current_status: str,
    next_status: str,
    role: str,
) -> bool:
    if current_status == next_status:
        return True
    if role == "admin":
        return True

    engineering_roles = {"engineer", "supervisor"}
    approval_roles = {"supervisor"}

    rules = {
        "Open": {
            "In Progress": engineering_roles,
            "Close Requested": engineering_roles,
            "Approved": approval_roles,
            "Closed": approval_roles,
            "Rejected": approval_roles,
        },
        "In Progress": {
            "Close Requested": engineering_roles,
            "Open": engineering_roles,
        },
        "Close Requested": {
            "Approved": approval_roles,
            "Closed": approval_roles,
            "Rejected": approval_roles,
            "In Progress": engineering_roles,
        },
        "Approved": {
            "Verified": approval_roles,
            "Closed": approval_roles,
            "In Progress": engineering_roles,
        },
        "Closed": {
            "Verified": approval_roles,
        },
        "Rejected": {
            "In Progress": engineering_roles,
            "Open": engineering_roles,
        },
        "Verified": {},
    }

    allowed = rules.get(current_status, {}).get(next_status, set())
    return role in allowed


def _apply_non_conformity_workflow_transition(
    row: NonConformity,
    next_status: str,
    user: CurrentUser,
    rejected_reason: Optional[str] = None,
) -> None:
    current_status = row.status
    if not _workflow_transition_allowed("nc", current_status, next_status, user.role):
        raise HTTPException(
            status_code=403,
            detail=f"Transition {current_status} -> {next_status} is not allowed for role '{user.role}'.",
        )

    now = datetime.utcnow()
    row.status = next_status

    if next_status == "Close Requested":
        row.close_requested_at = now
        row.close_requested_by = user.username
    if next_status in {"Approved", "Closed", "Verified"}:
        row.approved_at = row.approved_at or now
        row.approved_by = row.approved_by or user.username
        row.closed_at = now
        row.closed_by = user.username
    if next_status == "Verified":
        row.verified_at = now
        row.verified_by = user.username
    if next_status == "Rejected":
        row.rejected_at = now
        row.rejected_by = user.username
        row.rejected_reason = _clean_text(rejected_reason) or row.rejected_reason
        row.closed_at = None
        row.closed_by = None
    if next_status in {"Open", "In Progress"}:
        row.close_requested_at = None
        row.close_requested_by = None
        row.approved_at = None
        row.approved_by = None
        row.closed_at = None
        row.closed_by = None
        row.verified_at = None
        row.verified_by = None
        if current_status == "Rejected":
            row.rejected_at = None
            row.rejected_by = None
            row.rejected_reason = None


def _apply_capa_workflow_transition(
    row: CapaAction,
    next_status: str,
    user: CurrentUser,
    rejected_reason: Optional[str] = None,
) -> None:
    current_status = row.status
    if not _workflow_transition_allowed("capa", current_status, next_status, user.role):
        raise HTTPException(
            status_code=403,
            detail=f"Transition {current_status} -> {next_status} is not allowed for role '{user.role}'.",
        )

    now = datetime.utcnow()
    row.status = next_status

    if next_status == "Close Requested":
        row.close_requested_at = now
        row.close_requested_by = user.username
    if next_status in {"Approved", "Closed", "Verified"}:
        row.approved_at = row.approved_at or now
        row.approved_by = row.approved_by or user.username
        row.completed_at = now
        row.completed_by = user.username
    if next_status == "Verified":
        row.verified_at = now
        row.verified_by = user.username
    if next_status == "Rejected":
        row.rejected_at = now
        row.rejected_by = user.username
        row.rejected_reason = _clean_text(rejected_reason) or row.rejected_reason
        row.completed_at = None
        row.completed_by = None
    if next_status in {"Open", "In Progress"}:
        row.close_requested_at = None
        row.close_requested_by = None
        row.approved_at = None
        row.approved_by = None
        row.completed_at = None
        row.completed_by = None
        row.verified_at = None
        row.verified_by = None
        if current_status == "Rejected":
            row.rejected_at = None
            row.rejected_by = None
            row.rejected_reason = None


def _find_open_spc_non_conformity(
    session: Session,
    spec_id: uuid.UUID,
) -> Optional[NonConformity]:
    return session.exec(
        select(NonConformity).where(
            NonConformity.weight_specification_id == spec_id,
            NonConformity.source == "spc",
            NonConformity.status.in_(list(_open_workflow_statuses())),
        )
    ).first()


def _create_non_conformity_from_spc_chart(
    session: Session,
    spec: WeightSamplingSpec,
    chart: Dict[str, Any],
    detected_by: str,
    title_prefix: str = "Desvio SPC",
) -> NonConformity:
    latest_point = chart["points"][-1] if chart["points"] else None
    severity = "critical" if chart["alert_level"] == "critical" else "high"
    top_rule = chart["rules"]["violations"][0]["rule_code"] if chart["rules"]["violations"] else "NO_RULE"
    row = NonConformity(
        asset_id=spec.asset_id,
        product_reference_id=spec.product_reference_id,
        process_standard_id=spec.process_standard_id,
        weight_specification_id=spec.id,
        weight_sample_id=uuid.UUID(latest_point["sample_id"]) if latest_point else None,
        source="spc",
        severity=severity,
        status="Open",
        title=f"{title_prefix} - {spec.name}",
        description=f"Alerta SPC ({chart['alert_level']}) activada por regla {top_rule}.",
        detected_by=detected_by,
    )
    session.add(row)
    session.flush()
    return row


def _maybe_create_non_conformity_from_spc_event(
    session: Session,
    spec: WeightSamplingSpec,
    actor: str,
    limit: int = 300,
    minimum_alert_level: str = "warning",
) -> Optional[NonConformity]:
    chart = get_weight_spc_chart(
        spec_id=spec.id,
        limit=limit,
        include_rules=True,
        session=session,
    )
    alert_level = chart["alert_level"]
    if _minimum_alert_rank(alert_level) < _minimum_alert_rank(_normalize_alert_level(minimum_alert_level)):
        return None

    existing = _find_open_spc_non_conformity(session=session, spec_id=spec.id)
    if existing:
        return existing
    return _create_non_conformity_from_spc_chart(
        session=session,
        spec=spec,
        chart=chart,
        detected_by=actor,
    )


def _capability_run_payload(session: Session, row: WeightCapabilityRun) -> Dict[str, Any]:
    spec = session.get(WeightSamplingSpec, row.specification_id)
    action = session.get(ImprovementAction, row.improvement_action_id) if row.improvement_action_id else None
    return {
        "id": str(row.id),
        "specification_id": str(row.specification_id),
        "specification_name": spec.name if spec else None,
        "run_type": row.run_type,
        "triggered_by": row.triggered_by,
        "triggered_at": row.triggered_at.isoformat(),
        "limit_window": row.limit_window,
        "sample_count": row.sample_count,
        "center_line": row.center_line,
        "sigma_within": row.sigma_within,
        "sigma_overall": row.sigma_overall,
        "cp": row.cp,
        "cpk": row.cpk,
        "pp": row.pp,
        "ppk": row.ppk,
        "capability_status": row.capability_status,
        "alert_level": row.alert_level,
        "improvement_action_id": str(row.improvement_action_id) if row.improvement_action_id else None,
        "improvement_action_status": action.status if action else None,
        "notes": row.notes,
        "created_at": row.created_at.isoformat(),
    }


def _find_open_capability_action(session: Session, spec_id: uuid.UUID) -> Optional[ImprovementAction]:
    prefix = f"SPC-CAPABILITY:{spec_id}:"
    rows = session.exec(
        select(ImprovementAction).where(ImprovementAction.source_document.like(f"{prefix}%"))
    ).all()
    for row in rows:
        if row.status not in {"Closed", "Verified"}:
            return row
    return None


def _ensure_capability_improvement_action(
    session: Session,
    spec: WeightSamplingSpec,
    run: WeightCapabilityRun,
    capability_status: str,
    actor: str,
) -> Optional[ImprovementAction]:
    if capability_status not in {"marginal", "not_capable"}:
        return None

    existing = _find_open_capability_action(session=session, spec_id=spec.id)
    if existing:
        return existing

    due = date.today() + timedelta(days=14)
    severity_message = "marginal" if capability_status == "marginal" else "no capaz"
    action = ImprovementAction(
        asset_id=spec.asset_id,
        source_document=f"SPC-CAPABILITY:{spec.id}:{run.id}",
        description=f"[SPC] Proceso {severity_message} en {spec.name} (Cp/Cpk/Pp/Ppk).",
        responsible=actor,
        due_date=due,
        status="Open",
        completion_date=None,
    )
    session.add(action)
    session.flush()
    return action


def _create_capability_run(
    session: Session,
    spec: WeightSamplingSpec,
    limit: int,
    run_type: str,
    actor: str,
    auto_link_improvement_action: bool,
    notes: Optional[str],
) -> WeightCapabilityRun:
    normalized_run_type = (run_type or "on_demand").strip().lower()
    if normalized_run_type not in {"on_demand", "batch"}:
        normalized_run_type = "on_demand"

    capability_data = get_weight_spc_capability(spec_id=spec.id, limit=limit, session=session)
    chart_data = get_weight_spc_chart(spec_id=spec.id, limit=limit, include_rules=True, session=session)

    run = WeightCapabilityRun(
        specification_id=spec.id,
        run_type=normalized_run_type,
        triggered_by=actor,
        limit_window=limit,
        sample_count=capability_data["samples_count"],
        center_line=capability_data.get("process_mean"),
        sigma_within=capability_data.get("sigma_within"),
        sigma_overall=capability_data.get("sigma_overall"),
        cp=capability_data.get("cp"),
        cpk=capability_data.get("cpk"),
        pp=capability_data.get("pp"),
        ppk=capability_data.get("ppk"),
        capability_status=capability_data.get("status") or "insufficient_data",
        alert_level=chart_data.get("alert_level"),
        notes=_clean_text(notes),
    )
    session.add(run)
    session.flush()

    if auto_link_improvement_action:
        linked = _ensure_capability_improvement_action(
            session=session,
            spec=spec,
            run=run,
            capability_status=run.capability_status,
            actor=actor,
        )
        if linked:
            run.improvement_action_id = linked.id
            session.add(run)

    return run

def _non_conformity_payload(session: Session, row: NonConformity) -> Dict[str, Any]:
    asset = session.get(Asset, row.asset_id) if row.asset_id else None
    reference = session.get(ProductReference, row.product_reference_id) if row.product_reference_id else None
    standard = session.get(ProcessStandard, row.process_standard_id) if row.process_standard_id else None
    spec = session.get(WeightSamplingSpec, row.weight_specification_id) if row.weight_specification_id else None

    capa_count = len(
        session.exec(select(CapaAction).where(CapaAction.non_conformity_id == row.id)).all()
    )
    open_capa = len(
        [
            item
            for item in session.exec(
                select(CapaAction).where(CapaAction.non_conformity_id == row.id)
            ).all()
            if not _is_closed_status(item.status)
        ]
    )
    return {
        "id": str(row.id),
        "asset_id": str(row.asset_id) if row.asset_id else None,
        "asset_name": asset.name if asset else None,
        "product_reference_id": str(row.product_reference_id) if row.product_reference_id else None,
        "reference_code": reference.code if reference else None,
        "process_standard_id": str(row.process_standard_id) if row.process_standard_id else None,
        "standard_id": str(standard.id) if standard else None,
        "weight_specification_id": str(row.weight_specification_id) if row.weight_specification_id else None,
        "weight_specification_name": spec.name if spec else None,
        "weight_sample_id": str(row.weight_sample_id) if row.weight_sample_id else None,
        "source": row.source,
        "severity": row.severity,
        "status": row.status,
        "title": row.title,
        "description": row.description,
        "root_cause": row.root_cause,
        "containment": row.containment,
        "detected_at": row.detected_at.isoformat(),
        "detected_by": row.detected_by,
        "close_requested_at": row.close_requested_at.isoformat() if row.close_requested_at else None,
        "close_requested_by": row.close_requested_by,
        "approved_at": row.approved_at.isoformat() if row.approved_at else None,
        "approved_by": row.approved_by,
        "verified_at": row.verified_at.isoformat() if row.verified_at else None,
        "verified_by": row.verified_by,
        "rejected_at": row.rejected_at.isoformat() if row.rejected_at else None,
        "rejected_by": row.rejected_by,
        "rejected_reason": row.rejected_reason,
        "closed_at": row.closed_at.isoformat() if row.closed_at else None,
        "closed_by": row.closed_by,
        "created_at": row.created_at.isoformat(),
        "updated_at": row.updated_at.isoformat(),
        "capa_actions_count": capa_count,
        "open_capa_actions_count": open_capa,
    }


def _capa_payload(session: Session, row: CapaAction) -> Dict[str, Any]:
    return {
        "id": str(row.id),
        "non_conformity_id": str(row.non_conformity_id),
        "improvement_action_id": str(row.improvement_action_id) if row.improvement_action_id else None,
        "action_type": row.action_type,
        "title": row.title,
        "description": row.description,
        "responsible": row.responsible,
        "due_date": row.due_date.isoformat() if row.due_date else None,
        "status": row.status,
        "verification_notes": row.verification_notes,
        "close_requested_at": row.close_requested_at.isoformat() if row.close_requested_at else None,
        "close_requested_by": row.close_requested_by,
        "approved_at": row.approved_at.isoformat() if row.approved_at else None,
        "approved_by": row.approved_by,
        "verified_at": row.verified_at.isoformat() if row.verified_at else None,
        "verified_by": row.verified_by,
        "rejected_at": row.rejected_at.isoformat() if row.rejected_at else None,
        "rejected_by": row.rejected_by,
        "rejected_reason": row.rejected_reason,
        "completed_at": row.completed_at.isoformat() if row.completed_at else None,
        "completed_by": row.completed_by,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat(),
        "updated_at": row.updated_at.isoformat(),
    }


def _create_improvement_action_for_capa(
    non_conformity: NonConformity,
    capa_action: CapaAction,
    session: Session,
) -> ImprovementAction:
    source = f"CAPA:{non_conformity.id}:{capa_action.action_type}"
    action = ImprovementAction(
        asset_id=non_conformity.asset_id,
        source_document=source,
        description=f"[CAPA] {capa_action.title}",
        responsible=capa_action.responsible,
        due_date=capa_action.due_date,
        status=capa_action.status.replace(" ", ""),
        completion_date=None,
    )
    session.add(action)
    session.flush()
    return action


@router.post("/weight-specs")
def create_weight_spec(
    payload: WeightSpecCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    asset_id, reference_id, standard_id = _resolve_links(
        session,
        payload.asset_id,
        payload.product_reference_id,
        payload.process_standard_id,
    )
    lower_limit, upper_limit = _validate_limits(
        payload.lower_limit,
        payload.upper_limit,
        payload.target_weight,
    )
    row = WeightSamplingSpec(
        name=payload.name.strip(),
        asset_id=asset_id,
        product_reference_id=reference_id,
        process_standard_id=standard_id,
        unit=(payload.unit or "g").strip().lower(),
        lower_limit=lower_limit,
        target_weight=payload.target_weight,
        upper_limit=upper_limit,
        warning_band_pct=payload.warning_band_pct,
        sample_size=payload.sample_size,
        notes=_clean_text(payload.notes),
        is_active=payload.is_active,
        created_by=user.username,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return _spec_payload(session, row, include_summary=True)


@router.get("/weight-specs")
def list_weight_specs(
    asset_id: Optional[uuid.UUID] = None,
    product_reference_id: Optional[uuid.UUID] = None,
    process_standard_id: Optional[uuid.UUID] = None,
    is_active: Optional[bool] = None,
    session: Session = Depends(get_session),
):
    stmt = select(WeightSamplingSpec)
    if asset_id:
        stmt = stmt.where(WeightSamplingSpec.asset_id == asset_id)
    if product_reference_id:
        stmt = stmt.where(WeightSamplingSpec.product_reference_id == product_reference_id)
    if process_standard_id:
        stmt = stmt.where(WeightSamplingSpec.process_standard_id == process_standard_id)
    if is_active is not None:
        stmt = stmt.where(WeightSamplingSpec.is_active == is_active)
    rows = session.exec(stmt.order_by(WeightSamplingSpec.updated_at.desc())).all()
    return [_spec_payload(session, row, include_summary=False) for row in rows]


@router.get("/weight-specs/{spec_id}")
def get_weight_spec(spec_id: uuid.UUID, session: Session = Depends(get_session)):
    row = session.get(WeightSamplingSpec, spec_id)
    if not row:
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    return _spec_payload(session, row, include_summary=True)


@router.patch("/weight-specs/{spec_id}")
def update_weight_spec(
    spec_id: uuid.UUID,
    payload: WeightSpecUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    row = session.get(WeightSamplingSpec, spec_id)
    if not row:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    data = payload.model_dump(exclude_unset=True)
    asset_id = data.get("asset_id", row.asset_id)
    reference_id = data.get("product_reference_id", row.product_reference_id)
    standard_id = data.get("process_standard_id", row.process_standard_id)
    asset_id, reference_id, standard_id = _resolve_links(session, asset_id, reference_id, standard_id)

    lower_limit = data.get("lower_limit", row.lower_limit)
    upper_limit = data.get("upper_limit", row.upper_limit)
    target_weight = data.get("target_weight", row.target_weight)
    _validate_limits(lower_limit, upper_limit, target_weight)

    if "name" in data:
        row.name = data["name"].strip()
    if "unit" in data and data["unit"] is not None:
        row.unit = data["unit"].strip().lower()
    if "notes" in data:
        row.notes = _clean_text(data["notes"])
    if "is_active" in data:
        row.is_active = bool(data["is_active"])
    if "warning_band_pct" in data and data["warning_band_pct"] is not None:
        row.warning_band_pct = data["warning_band_pct"]
    if "sample_size" in data and data["sample_size"] is not None:
        row.sample_size = data["sample_size"]

    row.asset_id = asset_id
    row.product_reference_id = reference_id
    row.process_standard_id = standard_id
    row.lower_limit = lower_limit
    row.upper_limit = upper_limit
    row.target_weight = target_weight
    row.updated_at = datetime.utcnow()

    session.add(row)
    session.commit()
    session.refresh(row)
    return _spec_payload(session, row, include_summary=True)


@router.delete("/weight-specs/{spec_id}", status_code=204)
def delete_weight_spec(
    spec_id: uuid.UUID,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    row = session.get(WeightSamplingSpec, spec_id)
    if not row:
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    samples = session.exec(select(WeightSample).where(WeightSample.specification_id == spec_id)).all()
    for sample in samples:
        session.delete(sample)
    session.delete(row)
    session.commit()


@router.post("/weight-specs/{spec_id}/samples")
def create_weight_sample(
    spec_id: uuid.UUID,
    payload: WeightSampleCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor", "operator"])),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    measured_at = payload.measured_at or datetime.utcnow()
    measured_by = _clean_text(payload.measured_by) or user.username
    status_color = _status_for_value(spec, payload.measured_value)
    deviation = payload.measured_value - _target_for_spec(spec)

    row = WeightSample(
        specification_id=spec.id,
        measured_value=payload.measured_value,
        measured_at=measured_at,
        measured_by=measured_by,
        batch_code=_clean_text(payload.batch_code),
        shift=_clean_text(payload.shift),
        notes=_clean_text(payload.notes),
        deviation=round(deviation, 6),
        status_color=status_color,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    auto_nc = None
    if payload.auto_create_non_conformity:
        auto_nc = _maybe_create_non_conformity_from_spc_event(
            session=session,
            spec=spec,
            actor=user.username,
            minimum_alert_level=payload.minimum_alert_level,
        )
        if auto_nc:
            session.commit()
            session.refresh(auto_nc)
    response = _sample_payload(row)
    if auto_nc:
        response["auto_non_conformity_id"] = str(auto_nc.id)
    return response


@router.get("/weight-specs/{spec_id}/samples")
def list_weight_samples(
    spec_id: uuid.UUID,
    limit: int = Query(200, ge=1, le=2000),
    session: Session = Depends(get_session),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    rows = session.exec(
        select(WeightSample)
        .where(WeightSample.specification_id == spec_id)
        .order_by(WeightSample.measured_at.desc())  # type: ignore
        .limit(limit)
    ).all()
    return [_sample_payload(row) for row in rows]


@router.get("/weight-specs/{spec_id}/summary")
def get_weight_summary(spec_id: uuid.UUID, session: Session = Depends(get_session)):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    rows = session.exec(select(WeightSample).where(WeightSample.specification_id == spec_id)).all()
    return {
        "specification_id": str(spec.id),
        "specification_name": spec.name,
        "unit": spec.unit,
        **_build_summary(spec, rows),
    }


@router.get("/weight-specs/{spec_id}/spc/chart")
def get_weight_spc_chart(
    spec_id: uuid.UUID,
    limit: int = Query(200, ge=20, le=2000),
    include_rules: bool = True,
    session: Session = Depends(get_session),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    sample_rows_desc = session.exec(
        select(WeightSample)
        .where(WeightSample.specification_id == spec_id)
        .order_by(WeightSample.measured_at.desc())  # type: ignore
        .limit(limit)
    ).all()
    sample_rows = list(reversed(sample_rows_desc))
    values = [row.measured_value for row in sample_rows]
    stats = _compute_spc_stats(values)
    points = _build_spc_points(
        samples=sample_rows,
        center_line=stats["center_line"],
        sigma_within=stats["sigma_within"],
    )
    sample_ids = [str(row.id) for row in sample_rows]
    violations = (
        _detect_western_electric_rules(
            values=values,
            sample_ids=sample_ids,
            center_line=stats["center_line"],
            sigma=stats["sigma_within"],
        )
        if include_rules
        else []
    )
    alert_level = _spc_alert_level(violations)
    return {
        "specification_id": str(spec.id),
        "specification_name": spec.name,
        "unit": spec.unit,
        "chart_type": "I-MR",
        "samples_count": len(sample_rows),
        **stats,
        "points": points,
        "rules": {
            "rule_set": "western_electric",
            "violations_count": len(violations),
            "critical_count": sum(1 for row in violations if row["severity"] == "critical"),
            "warning_count": sum(1 for row in violations if row["severity"] == "warning"),
            "violations": violations,
        },
        "alert_level": alert_level,
    }


@router.get("/weight-specs/{spec_id}/spc/capability")
def get_weight_spc_capability(
    spec_id: uuid.UUID,
    limit: int = Query(300, ge=20, le=5000),
    session: Session = Depends(get_session),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    sample_rows_desc = session.exec(
        select(WeightSample)
        .where(WeightSample.specification_id == spec_id)
        .order_by(WeightSample.measured_at.desc())  # type: ignore
        .limit(limit)
    ).all()
    sample_rows = list(reversed(sample_rows_desc))
    values = [row.measured_value for row in sample_rows]
    stats = _compute_spc_stats(values)
    capability = _build_capability_payload(
        spec=spec,
        values=values,
        sigma_within=stats["sigma_within"],
        sigma_overall=stats["sigma_overall"],
    )
    return {
        "specification_id": str(spec.id),
        "specification_name": spec.name,
        "unit": spec.unit,
        "samples_count": len(values),
        "lower_limit": spec.lower_limit,
        "target_weight": _target_for_spec(spec),
        "upper_limit": spec.upper_limit,
        "sigma_within": stats["sigma_within"],
        "sigma_overall": stats["sigma_overall"],
        **capability,
    }


@router.post("/weight-specs/{spec_id}/spc/capability/runs")
def create_capability_run(
    spec_id: uuid.UUID,
    payload: CapabilityRunCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    run = _create_capability_run(
        session=session,
        spec=spec,
        limit=payload.limit,
        run_type=payload.run_type or "on_demand",
        actor=user.username,
        auto_link_improvement_action=payload.auto_link_improvement_action,
        notes=payload.notes,
    )
    session.commit()
    session.refresh(run)
    return _capability_run_payload(session, run)


@router.post("/spc/capability/runs/batch")
def create_capability_runs_batch(
    payload: CapabilityBatchCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    stmt = select(WeightSamplingSpec)
    if payload.asset_id:
        stmt = stmt.where(WeightSamplingSpec.asset_id == payload.asset_id)
    if payload.only_active:
        stmt = stmt.where(WeightSamplingSpec.is_active.is_(True))
    specs = session.exec(stmt).all()

    created_rows: List[Dict[str, Any]] = []
    for spec in specs:
        run = _create_capability_run(
            session=session,
            spec=spec,
            limit=payload.limit,
            run_type=payload.run_type or "batch",
            actor=user.username,
            auto_link_improvement_action=payload.auto_link_improvement_action,
            notes=payload.notes,
        )
        created_rows.append(_capability_run_payload(session, run))

    session.commit()
    return {
        "evaluated_specs": len(specs),
        "created_runs": len(created_rows),
        "runs": created_rows,
    }


@router.get("/weight-specs/{spec_id}/spc/capability/runs")
def list_capability_runs(
    spec_id: uuid.UUID,
    limit: int = Query(60, ge=1, le=500),
    session: Session = Depends(get_session),
):
    if not session.get(WeightSamplingSpec, spec_id):
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    rows = session.exec(
        select(WeightCapabilityRun)
        .where(WeightCapabilityRun.specification_id == spec_id)
        .order_by(WeightCapabilityRun.triggered_at.desc())  # type: ignore
        .limit(limit)
    ).all()
    return [_capability_run_payload(session, row) for row in rows]


@router.get("/weight-specs/{spec_id}/spc/capability/trend")
def get_capability_trend(
    spec_id: uuid.UUID,
    bucket: str = Query("month", pattern="^(day|week|month)$"),
    points: int = Query(12, ge=3, le=60),
    session: Session = Depends(get_session),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    rows = session.exec(
        select(WeightCapabilityRun)
        .where(WeightCapabilityRun.specification_id == spec_id)
        .order_by(WeightCapabilityRun.triggered_at.asc())  # type: ignore
    ).all()

    grouped: Dict[str, List[WeightCapabilityRun]] = {}
    for row in rows:
        dt = row.triggered_at
        if bucket == "day":
            key = dt.strftime("%Y-%m-%d")
        elif bucket == "week":
            iso = dt.isocalendar()
            key = f"{iso.year}-W{iso.week:02d}"
        else:
            key = dt.strftime("%Y-%m")
        grouped.setdefault(key, []).append(row)

    keys = sorted(grouped.keys())[-points:]
    series = []
    for key in keys:
        bucket_rows = grouped[key]
        cp_vals = [row.cp for row in bucket_rows if row.cp is not None]
        cpk_vals = [row.cpk for row in bucket_rows if row.cpk is not None]
        pp_vals = [row.pp for row in bucket_rows if row.pp is not None]
        ppk_vals = [row.ppk for row in bucket_rows if row.ppk is not None]
        status_counts: Dict[str, int] = {}
        for row in bucket_rows:
            status_counts[row.capability_status] = status_counts.get(row.capability_status, 0) + 1
        top_status = max(status_counts.items(), key=lambda pair: pair[1])[0] if status_counts else "insufficient_data"
        series.append(
            {
                "bucket": key,
                "runs": len(bucket_rows),
                "cp_avg": round(sum(cp_vals) / len(cp_vals), 6) if cp_vals else None,
                "cpk_avg": round(sum(cpk_vals) / len(cpk_vals), 6) if cpk_vals else None,
                "pp_avg": round(sum(pp_vals) / len(pp_vals), 6) if pp_vals else None,
                "ppk_avg": round(sum(ppk_vals) / len(ppk_vals), 6) if ppk_vals else None,
                "status": top_status,
                "status_counts": status_counts,
            }
        )

    return {
        "specification_id": str(spec.id),
        "specification_name": spec.name,
        "bucket": bucket,
        "points": points,
        "series": series,
    }


@router.get("/weight-specs/{spec_id}/spc/export/csv")
def export_weight_spc_csv(
    spec_id: uuid.UUID,
    limit: int = Query(300, ge=20, le=5000),
    session: Session = Depends(get_session),
):
    chart = get_weight_spc_chart(
        spec_id=spec_id,
        limit=limit,
        include_rules=True,
        session=session,
    )
    buffer = StringIO()
    buffer.write(
        "index,sample_id,measured_at,measured_value,moving_range,z_score,status_color,beyond_3sigma,measured_by,batch_code,shift,notes\n"
    )
    for point in chart["points"]:
        row = [
            str(point["index"]),
            point["sample_id"],
            point["measured_at"],
            str(point["measured_value"]),
            "" if point["moving_range"] is None else str(point["moving_range"]),
            "" if point["z_score"] is None else str(point["z_score"]),
            point["status_color"] or "",
            "1" if point["beyond_3sigma"] else "0",
            point.get("measured_by") or "",
            point.get("batch_code") or "",
            point.get("shift") or "",
            (point.get("notes") or "").replace("\n", " ").replace(",", ";"),
        ]
        buffer.write(",".join(row))
        buffer.write("\n")

    output = BytesIO(buffer.getvalue().encode("utf-8"))
    output.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="takta_spc_{spec_id}.csv"'
    }
    return StreamingResponse(output, media_type="text/csv; charset=utf-8", headers=headers)


class SpcNonConformityCreateRequest(BaseModel):
    limit: int = PydanticField(default=300, ge=20, le=2000)
    title_prefix: str = "Desvio SPC"


@router.post("/non-conformities/from-spc/{spec_id}")
def create_non_conformity_from_spc(
    spec_id: uuid.UUID,
    payload: SpcNonConformityCreateRequest,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    spec = session.get(WeightSamplingSpec, spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    chart = get_weight_spc_chart(
        spec_id=spec_id,
        limit=payload.limit,
        include_rules=True,
        session=session,
    )
    if chart["alert_level"] == "healthy":
        raise HTTPException(status_code=400, detail="SPC has no active alert for this specification.")

    existing = _find_open_spc_non_conformity(session=session, spec_id=spec_id)
    if existing:
        return _non_conformity_payload(session, existing)

    non_conformity = _create_non_conformity_from_spc_chart(
        session=session,
        spec=spec,
        chart=chart,
        detected_by=user.username,
        title_prefix=payload.title_prefix,
    )
    session.commit()
    session.refresh(non_conformity)
    return _non_conformity_payload(session, non_conformity)


@router.post("/non-conformities")
def create_non_conformity(
    payload: NonConformityCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    asset_id, reference_id, standard_id = _resolve_links(
        session,
        payload.asset_id,
        payload.product_reference_id,
        payload.process_standard_id,
    )
    if payload.weight_specification_id and not session.get(WeightSamplingSpec, payload.weight_specification_id):
        raise HTTPException(status_code=404, detail="Weight specification not found.")
    if payload.weight_sample_id and not session.get(WeightSample, payload.weight_sample_id):
        raise HTTPException(status_code=404, detail="Weight sample not found.")

    normalized_status = _normalize_workflow_status(payload.status, default="Open")
    row = NonConformity(
        asset_id=asset_id,
        product_reference_id=reference_id,
        process_standard_id=standard_id,
        weight_specification_id=payload.weight_specification_id,
        weight_sample_id=payload.weight_sample_id,
        source=(payload.source or "manual").strip().lower(),
        severity=_normalize_severity(payload.severity, default="medium"),
        status="Open",
        title=payload.title.strip(),
        description=_clean_text(payload.description),
        root_cause=_clean_text(payload.root_cause),
        containment=_clean_text(payload.containment),
        detected_by=user.username,
    )
    if normalized_status != "Open":
        _apply_non_conformity_workflow_transition(
            row=row,
            next_status=normalized_status,
            user=user,
        )
    session.add(row)
    session.commit()
    session.refresh(row)
    return _non_conformity_payload(session, row)


@router.post("/non-conformities/auto-generate")
def auto_generate_non_conformities_from_spc(
    payload: NonConformityAutoGenerateRequest,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    stmt = select(WeightSamplingSpec)
    if payload.only_active_specs:
        stmt = stmt.where(WeightSamplingSpec.is_active.is_(True))
    specs = session.exec(stmt).all()

    created: List[Dict[str, Any]] = []
    reused: List[Dict[str, Any]] = []
    for spec in specs:
        chart = get_weight_spc_chart(
            spec_id=spec.id,
            limit=payload.limit,
            include_rules=True,
            session=session,
        )
        alert_level = chart["alert_level"]
        if _minimum_alert_rank(alert_level) < _minimum_alert_rank(_normalize_alert_level(payload.minimum_alert_level)):
            continue
        existing = _find_open_spc_non_conformity(session=session, spec_id=spec.id)
        if existing:
            reused.append(_non_conformity_payload(session, existing))
            continue
        row = _create_non_conformity_from_spc_chart(
            session=session,
            spec=spec,
            chart=chart,
            detected_by=user.username,
        )
        created.append(_non_conformity_payload(session, row))

    session.commit()
    return {
        "evaluated_specs": len(specs),
        "created_count": len(created),
        "reused_open_count": len(reused),
        "created": created,
        "reused": reused,
    }


@router.get("/non-conformities")
def list_non_conformities(
    asset_id: Optional[uuid.UUID] = None,
    source: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
    weight_specification_id: Optional[uuid.UUID] = None,
    session: Session = Depends(get_session),
):
    stmt = select(NonConformity)
    if asset_id:
        stmt = stmt.where(NonConformity.asset_id == asset_id)
    if source:
        stmt = stmt.where(NonConformity.source == source.strip().lower())
    if severity:
        stmt = stmt.where(NonConformity.severity == _normalize_severity(severity, default="medium"))
    if status:
        stmt = stmt.where(NonConformity.status == _normalize_workflow_status(status))
    if weight_specification_id:
        stmt = stmt.where(NonConformity.weight_specification_id == weight_specification_id)
    rows = session.exec(stmt.order_by(NonConformity.detected_at.desc())).all()
    return [_non_conformity_payload(session, row) for row in rows]


@router.patch("/non-conformities/{non_conformity_id}")
def update_non_conformity(
    non_conformity_id: uuid.UUID,
    payload: NonConformityUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    row = session.get(NonConformity, non_conformity_id)
    if not row:
        raise HTTPException(status_code=404, detail="Non-conformity not found.")
    data = payload.model_dump(exclude_unset=True)
    if "severity" in data and data["severity"] is not None:
        row.severity = _normalize_severity(data["severity"], default=row.severity)
    if "status" in data and data["status"] is not None:
        next_status = _normalize_workflow_status(data["status"], default=row.status)
        _apply_non_conformity_workflow_transition(
            row=row,
            next_status=next_status,
            user=user,
            rejected_reason=data.get("rejected_reason"),
        )
    if "title" in data and data["title"] is not None:
        row.title = data["title"].strip()
    if "description" in data:
        row.description = _clean_text(data["description"])
    if "root_cause" in data:
        row.root_cause = _clean_text(data["root_cause"])
    if "containment" in data:
        row.containment = _clean_text(data["containment"])
    if "rejected_reason" in data and data["rejected_reason"] is not None:
        row.rejected_reason = _clean_text(data["rejected_reason"])
    row.updated_at = datetime.utcnow()
    session.add(row)
    session.commit()
    session.refresh(row)
    return _non_conformity_payload(session, row)


@router.delete("/non-conformities/{non_conformity_id}", status_code=204)
def delete_non_conformity(
    non_conformity_id: uuid.UUID,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    row = session.get(NonConformity, non_conformity_id)
    if not row:
        raise HTTPException(status_code=404, detail="Non-conformity not found.")
    capa_rows = session.exec(
        select(CapaAction).where(CapaAction.non_conformity_id == non_conformity_id)
    ).all()
    for capa in capa_rows:
        session.delete(capa)
    session.delete(row)
    session.commit()


@router.post("/non-conformities/{non_conformity_id}/capa-actions")
def create_capa_action(
    non_conformity_id: uuid.UUID,
    payload: CapaActionCreate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    non_conformity = session.get(NonConformity, non_conformity_id)
    if not non_conformity:
        raise HTTPException(status_code=404, detail="Non-conformity not found.")

    normalized_status = _normalize_workflow_status(payload.status, default="Open")
    row = CapaAction(
        non_conformity_id=non_conformity_id,
        action_type=(payload.action_type or "Corrective").strip().title(),
        title=payload.title.strip(),
        description=_clean_text(payload.description),
        responsible=payload.responsible.strip(),
        due_date=payload.due_date,
        status="Open",
        created_by=user.username,
    )
    session.add(row)
    session.flush()

    if normalized_status != "Open":
        _apply_capa_workflow_transition(
            row=row,
            next_status=normalized_status,
            user=user,
        )

    if payload.auto_link_improvement_action:
        linked = _create_improvement_action_for_capa(
            non_conformity=non_conformity,
            capa_action=row,
            session=session,
        )
        row.improvement_action_id = linked.id
        session.add(row)

    session.commit()
    session.refresh(row)
    return _capa_payload(session, row)


@router.get("/non-conformities/{non_conformity_id}/capa-actions")
def list_capa_actions(
    non_conformity_id: uuid.UUID,
    session: Session = Depends(get_session),
):
    if not session.get(NonConformity, non_conformity_id):
        raise HTTPException(status_code=404, detail="Non-conformity not found.")
    rows = session.exec(
        select(CapaAction)
        .where(CapaAction.non_conformity_id == non_conformity_id)
        .order_by(CapaAction.created_at.desc())  # type: ignore
    ).all()
    return [_capa_payload(session, row) for row in rows]


@router.patch("/capa-actions/{capa_action_id}")
def update_capa_action(
    capa_action_id: uuid.UUID,
    payload: CapaActionUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    row = session.get(CapaAction, capa_action_id)
    if not row:
        raise HTTPException(status_code=404, detail="CAPA action not found.")
    data = payload.model_dump(exclude_unset=True)
    if "action_type" in data and data["action_type"] is not None:
        row.action_type = data["action_type"].strip().title()
    if "title" in data and data["title"] is not None:
        row.title = data["title"].strip()
    if "description" in data:
        row.description = _clean_text(data["description"])
    if "responsible" in data and data["responsible"] is not None:
        row.responsible = data["responsible"].strip()
    if "due_date" in data:
        row.due_date = data["due_date"]
    if "verification_notes" in data:
        row.verification_notes = _clean_text(data["verification_notes"])
    if "status" in data and data["status"] is not None:
        next_status = _normalize_workflow_status(data["status"], default=row.status)
        _apply_capa_workflow_transition(
            row=row,
            next_status=next_status,
            user=user,
            rejected_reason=data.get("rejected_reason"),
        )
    if "rejected_reason" in data and data["rejected_reason"] is not None:
        row.rejected_reason = _clean_text(data["rejected_reason"])
    row.updated_at = datetime.utcnow()
    session.add(row)

    if row.improvement_action_id:
        linked = session.get(ImprovementAction, row.improvement_action_id)
        if linked:
            linked.status = row.status.replace(" ", "")
            linked.responsible = row.responsible
            linked.due_date = row.due_date
            if _is_closed_status(row.status):
                linked.completion_date = date.today()
            else:
                linked.completion_date = None
            session.add(linked)

    session.commit()
    session.refresh(row)
    return _capa_payload(session, row)


@router.delete("/capa-actions/{capa_action_id}", status_code=204)
def delete_capa_action(
    capa_action_id: uuid.UUID,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer"])),
):
    row = session.get(CapaAction, capa_action_id)
    if not row:
        raise HTTPException(status_code=404, detail="CAPA action not found.")
    session.delete(row)
    session.commit()


@router.get("/capa/dashboard")
def capa_dashboard(session: Session = Depends(get_session)):
    nc_rows = session.exec(select(NonConformity)).all()
    capa_rows = session.exec(select(CapaAction)).all()
    today = date.today()
    open_nc = [row for row in nc_rows if not _is_closed_status(row.status)]
    open_capa = [row for row in capa_rows if not _is_closed_status(row.status)]
    overdue_capa = [
        row for row in open_capa if row.due_date and row.due_date < today
    ]
    critical_nc = [row for row in open_nc if row.severity == "critical"]
    linked_actions = [row for row in capa_rows if row.improvement_action_id is not None]
    return {
        "non_conformities_total": len(nc_rows),
        "non_conformities_open": len(open_nc),
        "non_conformities_critical_open": len(critical_nc),
        "capa_actions_total": len(capa_rows),
        "capa_actions_open": len(open_capa),
        "capa_actions_overdue": len(overdue_capa),
        "capa_actions_linked_to_ci": len(linked_actions),
    }


@router.patch("/weight-samples/{sample_id}")
def update_weight_sample(
    sample_id: uuid.UUID,
    payload: WeightSampleUpdate,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor", "operator"])),
):
    row = session.get(WeightSample, sample_id)
    if not row:
        raise HTTPException(status_code=404, detail="Weight sample not found.")
    spec = session.get(WeightSamplingSpec, row.specification_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Weight specification not found.")

    data = payload.model_dump(exclude_unset=True)
    if "measured_value" in data and data["measured_value"] is not None:
        row.measured_value = data["measured_value"]
    if "measured_at" in data and data["measured_at"] is not None:
        row.measured_at = data["measured_at"]
    if "measured_by" in data:
        row.measured_by = _clean_text(data["measured_by"]) or row.measured_by
    if "batch_code" in data:
        row.batch_code = _clean_text(data["batch_code"])
    if "shift" in data:
        row.shift = _clean_text(data["shift"])
    if "notes" in data:
        row.notes = _clean_text(data["notes"])

    row.status_color = _status_for_value(spec, row.measured_value)
    row.deviation = round(row.measured_value - _target_for_spec(spec), 6)
    session.add(row)
    session.commit()
    session.refresh(row)
    auto_create_nc = data.get("auto_create_non_conformity", True)
    minimum_alert_level = data.get("minimum_alert_level", "warning")
    auto_nc = None
    if auto_create_nc:
        auto_nc = _maybe_create_non_conformity_from_spc_event(
            session=session,
            spec=spec,
            actor=user.username,
            minimum_alert_level=minimum_alert_level,
        )
        if auto_nc:
            session.commit()
            session.refresh(auto_nc)
    response = _sample_payload(row)
    if auto_nc:
        response["auto_non_conformity_id"] = str(auto_nc.id)
    return response


@router.delete("/weight-samples/{sample_id}", status_code=204)
def delete_weight_sample(
    sample_id: uuid.UUID,
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    row = session.get(WeightSample, sample_id)
    if not row:
        raise HTTPException(status_code=404, detail="Weight sample not found.")
    session.delete(row)
    session.commit()


def _build_template_workbook(entity: str, session: Session) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active

    if entity == "specs":
        sheet.title = "specs"
        sheet.append(
            [
                "id",
                "name",
                "asset_id",
                "product_reference_id",
                "process_standard_id",
                "unit",
                "lower_limit",
                "target_weight",
                "upper_limit",
                "warning_band_pct",
                "sample_size",
                "notes",
                "is_active",
            ]
        )
        sheet.append(
            [
                "",
                "Control Peso SKU-001",
                "",
                "",
                "",
                "g",
                95.0,
                100.0,
                105.0,
                0.1,
                5,
                "Especificacion base de ejemplo",
                True,
            ]
        )
        catalog = workbook.create_sheet("catalog")
        catalog.append(["type", "id", "name", "extra"])
        for asset in session.exec(select(Asset)).all():
            catalog.append(["asset", str(asset.id), asset.name, asset.type])
        for reference in session.exec(select(ProductReference)).all():
            catalog.append(["reference", str(reference.id), reference.code, reference.description])
        for standard in session.exec(select(ProcessStandard)).all():
            activity = session.get(StandardActivity, standard.activity_id)
            extra = activity.name if activity else ""
            catalog.append(["standard", str(standard.id), str(standard.asset_id), extra])
    else:
        sheet.title = "samples"
        sheet.append(
            [
                "id",
                "specification_id",
                "measured_value",
                "measured_at",
                "measured_by",
                "batch_code",
                "shift",
                "notes",
            ]
        )
        first_spec = session.exec(select(WeightSamplingSpec).order_by(WeightSamplingSpec.created_at.desc())).first()
        sheet.append(
            [
                "",
                str(first_spec.id) if first_spec else "",
                100.2,
                datetime.utcnow().isoformat(timespec="seconds"),
                "qa.user",
                "Lote-001",
                "Manana",
                "Muestra inicial",
            ]
        )
        specs = workbook.create_sheet("specs_catalog")
        specs.append(["id", "name", "unit", "lower_limit", "target_weight", "upper_limit"])
        for row in session.exec(select(WeightSamplingSpec)).all():
            specs.append([str(row.id), row.name, row.unit, row.lower_limit, row.target_weight, row.upper_limit])

    return workbook


def _build_export_workbook(entity: str, session: Session) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active

    if entity == "specs":
        sheet.title = "specs"
        sheet.append(
            [
                "id",
                "name",
                "asset_id",
                "product_reference_id",
                "process_standard_id",
                "unit",
                "lower_limit",
                "target_weight",
                "upper_limit",
                "warning_band_pct",
                "sample_size",
                "notes",
                "is_active",
                "created_by",
                "created_at",
                "updated_at",
            ]
        )
        for row in session.exec(select(WeightSamplingSpec)).all():
            sheet.append(
                [
                    str(row.id),
                    row.name,
                    str(row.asset_id) if row.asset_id else None,
                    str(row.product_reference_id) if row.product_reference_id else None,
                    str(row.process_standard_id) if row.process_standard_id else None,
                    row.unit,
                    row.lower_limit,
                    row.target_weight,
                    row.upper_limit,
                    row.warning_band_pct,
                    row.sample_size,
                    row.notes,
                    row.is_active,
                    row.created_by,
                    row.created_at.isoformat(),
                    row.updated_at.isoformat(),
                ]
            )
    else:
        sheet.title = "samples"
        sheet.append(
            [
                "id",
                "specification_id",
                "specification_name",
                "measured_value",
                "measured_at",
                "measured_by",
                "batch_code",
                "shift",
                "deviation",
                "status_color",
                "notes",
            ]
        )
        for row in session.exec(select(WeightSample).order_by(WeightSample.measured_at.desc())).all():
            spec = session.get(WeightSamplingSpec, row.specification_id)
            sheet.append(
                [
                    str(row.id),
                    str(row.specification_id),
                    spec.name if spec else None,
                    row.measured_value,
                    row.measured_at.isoformat(),
                    row.measured_by,
                    row.batch_code,
                    row.shift,
                    row.deviation,
                    row.status_color,
                    row.notes,
                ]
            )

    return workbook


def _import_specs(rows: List[tuple], session: Session, user: CurrentUser) -> tuple[int, int, List[str]]:
    created = 0
    updated = 0
    errors: List[str] = []

    for row_idx, row in enumerate(rows, start=2):
        row_values = list(row or ())
        if not any(value not in (None, "") for value in row_values):
            continue

        raw_id = row_values[0] if len(row_values) > 0 else None
        raw_name = _clean_text(row_values[1] if len(row_values) > 1 else None)
        raw_asset = row_values[2] if len(row_values) > 2 else None
        raw_reference = row_values[3] if len(row_values) > 3 else None
        raw_standard = row_values[4] if len(row_values) > 4 else None
        unit = _clean_text(row_values[5] if len(row_values) > 5 else None) or "g"
        lower_limit = _as_float(row_values[6] if len(row_values) > 6 else None)
        target_weight = _as_float(row_values[7] if len(row_values) > 7 else None)
        upper_limit = _as_float(row_values[8] if len(row_values) > 8 else None)
        warning_band_pct = _as_float(row_values[9] if len(row_values) > 9 else None)
        sample_size = _as_int(row_values[10] if len(row_values) > 10 else None)
        notes = _clean_text(row_values[11] if len(row_values) > 11 else None)
        is_active = _as_bool(row_values[12] if len(row_values) > 12 else True, default=True)

        if not raw_name:
            errors.append(f"Row {row_idx}: name is required.")
            continue
        if lower_limit is None or upper_limit is None:
            errors.append(f"Row {row_idx}: lower_limit and upper_limit are required.")
            continue

        try:
            spec_id = _to_uuid(raw_id, "id")
            asset_id = _to_uuid(raw_asset, "asset_id")
            reference_id = _to_uuid(raw_reference, "product_reference_id")
            standard_id = _to_uuid(raw_standard, "process_standard_id")
            asset_id, reference_id, standard_id = _resolve_links(session, asset_id, reference_id, standard_id)
            _validate_limits(lower_limit, upper_limit, target_weight)
        except HTTPException as exc:
            errors.append(f"Row {row_idx}: {exc.detail}")
            continue

        existing = session.get(WeightSamplingSpec, spec_id) if spec_id else None
        if spec_id and not existing:
            errors.append(f"Row {row_idx}: specification id not found.")
            continue

        if existing:
            existing.name = raw_name
            existing.asset_id = asset_id
            existing.product_reference_id = reference_id
            existing.process_standard_id = standard_id
            existing.unit = unit.lower()
            existing.lower_limit = lower_limit
            existing.target_weight = target_weight
            existing.upper_limit = upper_limit
            existing.warning_band_pct = warning_band_pct if warning_band_pct is not None else existing.warning_band_pct
            existing.sample_size = sample_size if sample_size is not None else existing.sample_size
            existing.notes = notes
            existing.is_active = is_active
            existing.updated_at = datetime.utcnow()
            session.add(existing)
            updated += 1
        else:
            session.add(
                WeightSamplingSpec(
                    name=raw_name,
                    asset_id=asset_id,
                    product_reference_id=reference_id,
                    process_standard_id=standard_id,
                    unit=unit.lower(),
                    lower_limit=lower_limit,
                    target_weight=target_weight,
                    upper_limit=upper_limit,
                    warning_band_pct=warning_band_pct if warning_band_pct is not None else 0.1,
                    sample_size=sample_size if sample_size is not None else 5,
                    notes=notes,
                    is_active=is_active,
                    created_by=user.username,
                )
            )
            created += 1

    return created, updated, errors


def _import_samples(rows: List[tuple], session: Session, user: CurrentUser) -> tuple[int, int, List[str]]:
    created = 0
    updated = 0
    errors: List[str] = []

    for row_idx, row in enumerate(rows, start=2):
        row_values = list(row or ())
        if not any(value not in (None, "") for value in row_values):
            continue

        raw_id = row_values[0] if len(row_values) > 0 else None
        raw_spec_id = row_values[1] if len(row_values) > 1 else None
        measured_value = _as_float(row_values[2] if len(row_values) > 2 else None)
        measured_at = _parse_datetime_cell(row_values[3] if len(row_values) > 3 else None)
        measured_by = _clean_text(row_values[4] if len(row_values) > 4 else None) or user.username
        batch_code = _clean_text(row_values[5] if len(row_values) > 5 else None)
        shift = _clean_text(row_values[6] if len(row_values) > 6 else None)
        notes = _clean_text(row_values[7] if len(row_values) > 7 else None)

        if measured_value is None:
            errors.append(f"Row {row_idx}: measured_value is required.")
            continue

        try:
            sample_id = _to_uuid(raw_id, "id")
            spec_id = _to_uuid(raw_spec_id, "specification_id")
            if not spec_id:
                raise HTTPException(status_code=400, detail="specification_id is required.")
        except HTTPException as exc:
            errors.append(f"Row {row_idx}: {exc.detail}")
            continue

        spec = session.get(WeightSamplingSpec, spec_id)
        if not spec:
            errors.append(f"Row {row_idx}: specification not found.")
            continue

        status_color = _status_for_value(spec, measured_value)
        deviation = round(measured_value - _target_for_spec(spec), 6)

        existing = session.get(WeightSample, sample_id) if sample_id else None
        if sample_id and not existing:
            errors.append(f"Row {row_idx}: sample id not found.")
            continue

        if existing:
            existing.specification_id = spec.id
            existing.measured_value = measured_value
            existing.measured_at = measured_at
            existing.measured_by = measured_by
            existing.batch_code = batch_code
            existing.shift = shift
            existing.notes = notes
            existing.deviation = deviation
            existing.status_color = status_color
            session.add(existing)
            updated += 1
        else:
            session.add(
                WeightSample(
                    specification_id=spec.id,
                    measured_value=measured_value,
                    measured_at=measured_at,
                    measured_by=measured_by,
                    batch_code=batch_code,
                    shift=shift,
                    notes=notes,
                    deviation=deviation,
                    status_color=status_color,
                )
            )
            created += 1

    return created, updated, errors


@router.get("/weights/xlsx/template")
def download_quality_template(
    entity: str = Query(..., pattern="^(specs|samples)$"),
    session: Session = Depends(get_session),
):
    workbook = _build_template_workbook(entity, session)
    return _workbook_response(workbook, f"takta_weight_{entity}_template.xlsx")


@router.get("/weights/xlsx/export")
def download_quality_export(
    entity: str = Query(..., pattern="^(specs|samples)$"),
    session: Session = Depends(get_session),
):
    workbook = _build_export_workbook(entity, session)
    return _workbook_response(workbook, f"takta_weight_{entity}_export.xlsx")


@router.post("/weights/xlsx/import")
def import_quality_xlsx(
    entity: str = Query(..., pattern="^(specs|samples)$"),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role(["admin", "engineer", "supervisor"])),
):
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file.")

    try:
        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLSX file.") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if entity == "specs":
        created, updated, errors = _import_specs(rows, session, user)
    else:
        created, updated, errors = _import_samples(rows, session, user)

    session.commit()
    return {
        "entity": entity,
        "created": created,
        "updated": updated,
        "errors_count": len(errors),
        "errors": errors,
    }
