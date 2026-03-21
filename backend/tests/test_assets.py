"""
Tests for /api/assets endpoints.
Covers: auth, CRUD, tree, breadcrumbs, cycle detection, orphan prevention.
"""
from io import BytesIO

from openpyxl import load_workbook


def test_root(client):
    """Health check."""
    response = client.get("/")
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "OK"
    assert "version" in data


def test_assets_requires_auth(client):
    response = client.get("/api/assets/")
    assert response.status_code == 401


# --- CREATE ---
def test_create_asset(client, auth_headers):
    response = client.post(
        "/api/assets/",
        json={
            "name": "Sede Pereira",
            "type": "Sede",
            "description": "Planta principal"
        },
        headers=auth_headers,
    )
    assert response.status_code == 201
    data = response.json()
    assert data["name"] == "Sede Pereira"
    assert data["type"] == "Sede"
    assert data["parent_id"] is None


def test_create_child_asset(client, auth_headers):
    parent = client.post(
        "/api/assets/",
        json={"name": "Sede Pereira", "type": "Sede"},
        headers=auth_headers,
    ).json()

    child = client.post(
        "/api/assets/",
        json={
            "name": "Planta Beneficio",
            "type": "Planta",
            "parent_id": parent["id"]
        },
        headers=auth_headers,
    )
    assert child.status_code == 201
    assert child.json()["parent_id"] == parent["id"]


def test_create_asset_invalid_parent(client, auth_headers):
    response = client.post(
        "/api/assets/",
        json={
            "name": "Orphan",
            "type": "Planta",
            "parent_id": "00000000-0000-0000-0000-000000000000"
        },
        headers=auth_headers,
    )
    assert response.status_code == 404


# --- READ ---
def test_list_assets(client, auth_headers):
    response = client.get("/api/assets/", headers=auth_headers)
    assert response.status_code == 200
    assert isinstance(response.json(), list)


def test_list_assets_filter_by_type(client, auth_headers):
    client.post("/api/assets/", json={"name": "Sede", "type": "Sede"}, headers=auth_headers)
    client.post("/api/assets/", json={"name": "Maquina", "type": "Maquina"}, headers=auth_headers)

    response = client.get("/api/assets/?type=Sede", headers=auth_headers)
    data = response.json()
    assert all(a["type"] == "Sede" for a in data)


def test_get_asset_by_id(client, auth_headers):
    created = client.post(
        "/api/assets/",
        json={"name": "Sede Pereira", "type": "Sede"},
        headers=auth_headers,
    ).json()

    response = client.get(f"/api/assets/{created['id']}", headers=auth_headers)
    assert response.status_code == 200
    assert response.json()["name"] == "Sede Pereira"


def test_get_asset_not_found(client, auth_headers):
    response = client.get("/api/assets/00000000-0000-0000-0000-000000000000", headers=auth_headers)
    assert response.status_code == 404


# --- TREE ---
def test_asset_tree_nested(client, auth_headers):
    sede = client.post("/api/assets/", json={"name": "Sede", "type": "Sede"}, headers=auth_headers).json()
    planta = client.post(
        "/api/assets/",
        json={"name": "Planta", "type": "Planta", "parent_id": sede["id"]},
        headers=auth_headers,
    ).json()
    client.post(
        "/api/assets/",
        json={"name": "Linea 1", "type": "Linea", "parent_id": planta["id"]},
        headers=auth_headers,
    )

    tree = client.get("/api/assets/tree", headers=auth_headers).json()
    assert len(tree) == 1
    assert tree[0]["name"] == "Sede"
    assert tree[0]["children"][0]["name"] == "Planta"
    assert tree[0]["children"][0]["children"][0]["name"] == "Linea 1"


# --- BREADCRUMBS ---
def test_breadcrumbs(client, auth_headers):
    sede = client.post("/api/assets/", json={"name": "Sede", "type": "Sede"}, headers=auth_headers).json()
    planta = client.post(
        "/api/assets/",
        json={"name": "Planta", "type": "Planta", "parent_id": sede["id"]},
        headers=auth_headers,
    ).json()
    linea = client.post(
        "/api/assets/",
        json={"name": "Linea 1", "type": "Linea", "parent_id": planta["id"]},
        headers=auth_headers,
    ).json()

    ctx = client.get(f"/api/assets/{linea['id']}/context", headers=auth_headers).json()
    assert ctx["depth"] == 3
    assert ctx["breadcrumbs"][0]["name"] == "Sede"
    assert ctx["breadcrumbs"][1]["name"] == "Planta"
    assert ctx["breadcrumbs"][2]["name"] == "Linea 1"


# --- DELETE ---
def test_delete_leaf_asset(client, auth_headers):
    created = client.post("/api/assets/", json={"name": "Temp", "type": "Sede"}, headers=auth_headers).json()
    response = client.delete(f"/api/assets/{created['id']}", headers=auth_headers)
    assert response.status_code == 204


def test_delete_asset_with_children_fails(client, auth_headers):
    parent = client.post("/api/assets/", json={"name": "Parent", "type": "Sede"}, headers=auth_headers).json()
    client.post(
        "/api/assets/",
        json={"name": "Child", "type": "Planta", "parent_id": parent["id"]},
        headers=auth_headers,
    )

    response = client.delete(f"/api/assets/{parent['id']}", headers=auth_headers)
    assert response.status_code == 400
    assert "children" in response.json()["detail"]


def test_assets_xlsx_template_and_export(client, auth_headers):
    template = client.get("/api/assets/xlsx/template", headers=auth_headers)
    assert template.status_code == 200

    template_wb = load_workbook(filename=BytesIO(template.content), data_only=True)
    assert "assets" in template_wb.sheetnames
    header = [cell.value for cell in next(template_wb["assets"].iter_rows(min_row=1, max_row=1))]
    assert header == ["asset_id", "name", "type", "description", "code", "parent_code"]

    client.post("/api/assets/", json={"name": "Sede QA", "type": "Sede"}, headers=auth_headers)
    exported = client.get("/api/assets/xlsx/export", headers=auth_headers)
    assert exported.status_code == 200
    export_wb = load_workbook(filename=BytesIO(exported.content), data_only=True)
    assert export_wb["assets"].max_row >= 2


def test_assets_xlsx_import(client, auth_headers):
    workbook = load_workbook(filename=BytesIO(client.get("/api/assets/xlsx/template", headers=auth_headers).content))
    sheet = workbook["assets"]

    # Keep headers, then inject our own rows
    for idx in range(sheet.max_row, 1, -1):
        sheet.delete_rows(idx, 1)

    sheet.append(["", "Sede Import", "Sede", "Nodo raiz importado", "SEDE-IMP", ""])
    sheet.append(["", "Linea Import", "Linea", "Nodo hijo importado", "LINEA-IMP", "SEDE-IMP"])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = client.post(
        "/api/assets/xlsx/import",
        headers=auth_headers,
        files={"file": ("assets_import.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["created"] == 2

    assets = client.get("/api/assets", headers=auth_headers).json()
    imported = {asset["name"]: asset for asset in assets}
    assert "Sede Import" in imported
    assert "Linea Import" in imported
    assert imported["Linea Import"]["parent_id"] == imported["Sede Import"]["id"]
